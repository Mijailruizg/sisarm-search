import openpyxl
import re
import unicodedata
import tempfile
import os
from .models import PartidaArancelaria, ImportLog


def _normalize(s: str) -> str:
    s = (s or '')
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r'[^0-9a-z]+', ' ', s)
    s = ' '.join(s.split())
    return s


def _split_normalize_ace22(full_ace22: str):
    if not full_ace22:
        return '', ''
    s = str(full_ace22).strip()
    sep_candidates = [';', '|', '/', ',', '\n']
    split_parts = None
    for sep in sep_candidates:
        if sep in s:
            split_parts = [p.strip() for p in s.split(sep) if p.strip()]
            break
    if not split_parts:
        if '  ' in s:
            split_parts = [p.strip() for p in s.split('  ') if p.strip()]
    if not split_parts:
        parts = [s]
    else:
        parts = split_parts

    bad_set = {'n', 'y', 's', 'si', 'no', '0', '1', 'yes', 'a'}
    def clean(val):
        if not val:
            return ''
        v = val.strip()
        if len(v) <= 1 and v.lower() in bad_set:
            return ''
        if v.lower() in bad_set:
            return ''
        return v

    chi = clean(parts[0]) if len(parts) >= 1 else ''
    prot = clean(parts[1]) if len(parts) >= 2 else ''
    return chi, prot


def _detect_chapter_from_code(codigo):
    """Detecta el número de capítulo (01-99) del prefijo del código.
    Ej: '0101XXXX' -> '01', '0301XXXX' -> '03', 'SIN02XX' -> None
    """
    if not codigo:
        return None
    s = str(codigo).strip()
    if len(s) >= 2 and s[:2].isdigit():
        return s[:2]
    return None


def _detect_headers(sheet):
    headers = []
    found = {}
    for cell in sheet[1]:
        val = cell.value
        headers.append(val)

    if not headers:
        return found

    header_normalized = [_normalize(h) for h in headers]

    patterns = {
        'codigo': r'(?:cod|codigo|code|hs|sa|arancelaria)',
        'descripcion': r'(?:desc|descripcion|description)',
        'partida': r'(?:partida|partition)',
        'capitulo': r'(?:cap|capitulo|chapter)',
        'gravamen': r'(?:grav|gravamen)',
        'ice_iehd': r'(?:ice|iehd)',
        'unidad_medida': r'(?:unidad|medida|unit)',
        'despacho_frontera': r'(?:despacho|frontera)',
        'tipo_documento': r'(?:tipo|documento|doc)',
        'entidad_emite': r'(?:entidad|emite)',
        'disp_legal': r'(?:disp|legal)',
        'can_ace36_ace47_ven': r'(?:ace36|ace47|ven)',
        'ace66_mexico': r'(?:ace66|mexico)',
        'subpartida': r'(?:subpartida)',
        'ace22_chi_prot': r'(?:ace22|chi_prot)',
        'ace22_chi': r'(?:chi)',
        'ace22_prot': r'(?:prot)',
    }

    for key, pattern in patterns.items():
        for idx, hn in enumerate(header_normalized):
            if re.search(pattern, hn):
                found[key] = idx
                break


    if 'codigo' not in found and 'partida' in found:
        found['codigo'] = found['partida']
    if 'codigo' in found and 'partida' not in found:
        found['partida'] = found['codigo']

    ace22_idx = found.get('ace22_chi_prot')
    ace22_chi_idx = found.get('ace22_chi')
    ace22_prot_idx = found.get('ace22_prot')

    if ace22_idx is None and (ace22_chi_idx is not None or ace22_prot_idx is not None):
        if ace22_chi_idx is not None:
            found['ace22_chi_prot'] = ace22_chi_idx

    if ace22_idx is not None:
        found['ace22_chi_prot'] = ace22_idx
    if ace22_chi_idx is not None:
        found['ace22_chi'] = ace22_chi_idx
    if ace22_prot_idx is not None:
        found['ace22_prot'] = ace22_prot_idx

    return found


def preview_import(source, update_existing=False):
    """Devuelve una estructura con filas detectadas y errores por fila para mostrar en preview.
    Detecta capítulos automáticamente por prefijo de código (01, 02, 03, etc).
    """
    if isinstance(source, str):
        wb = openpyxl.load_workbook(source)
    else:
        wb = openpyxl.load_workbook(source)
    sheet = wb.active
    found = _detect_headers(sheet)

    rows = []
    seen_codes = set()
    chapters_detected = {}
    line_no = 1
    for fila in sheet.iter_rows(min_row=2):
        line_no += 1
        row_vals = [cell.value for cell in fila]

        def get_by_key(key):
            idx = found.get(key)
            if idx is None or idx >= len(row_vals):
                return None
            return row_vals[idx]

        def limpiar(valor):
            return '' if valor in (None, '-', '–') else str(valor).strip()

        codigo = limpiar(get_by_key('codigo'))
        descripcion = limpiar(get_by_key('descripcion'))
        partida = limpiar(get_by_key('partida'))


        chapter_from_code = _detect_chapter_from_code(codigo)
        
        ace22_val = ''
        combined = get_by_key('ace22_chi_prot')
        chi_val = get_by_key('ace22_chi')
        prot_val = get_by_key('ace22_prot')
        if combined:
            ace22_val = limpiar(combined)
        else:
            chi = limpiar(chi_val)
            prot = limpiar(prot_val)
            if chi and prot:
                ace22_val = f"{chi}; {prot}"
            elif chi:
                ace22_val = chi
            elif prot:
                ace22_val = prot

        errors = []
        if not codigo:
            errors.append('codigo vacío')
        if not descripcion:
            errors.append('descripcion vacía')
        if not partida:
            errors.append('partida vacío')
        if not chapter_from_code:
            errors.append('codigo sin prefijo válido (debe empezar con 2 dígitos: 01, 02, etc)')

        if codigo:
            if codigo in seen_codes:
                errors.append('codigo duplicado en archivo')
            seen_codes.add(codigo)
            if not update_existing and PartidaArancelaria.objects.filter(codigo=codigo).exists():
                errors.append('codigo ya existe en la base de datos')

        if chapter_from_code:
            chapters_detected[chapter_from_code] = chapters_detected.get(chapter_from_code, 0) + 1

        if codigo or descripcion or partida or ace22_val or errors:
            rows.append({
                'line': line_no, 
                'codigo': codigo, 
                'descripcion': descripcion, 
                'chapter': chapter_from_code,
                'partida': partida, 
                'ace22': ace22_val, 
                'errors': errors
            })

    total = len(rows)
    errors_count = sum(1 for r in rows if r['errors'])
    chapters_list = sorted(list(chapters_detected.keys()))
    return {
        'rows': rows, 
        'total': total, 
        'errors_count': errors_count, 
        'chapters': chapters_list,
        'chapters_detail': chapters_detected
    }


def process_import(source, usuario=None, update_existing=False, sync_catalog=False, nombre_archivo=None):
    """Procesa importación con sincronización POR CAPÍTULO (múltiples capítulos).
    Detecta capítulos automáticamente por prefijo de código (01, 02, 03, 04, etc).
    Agrupa por capítulo y procesa cada uno independientemente.
    Si sync_catalog=True: para CADA capítulo del Excel, sincroniza solo ese capítulo
    """
    if isinstance(source, str):
        wb = openpyxl.load_workbook(source)
    else:
        wb = openpyxl.load_workbook(source)
    sheet = wb.active
    found = _detect_headers(sheet)

    total = 0
    errors = []
    rows_by_chapter = {} 
    codes_in_file = set()

    for fila in sheet.iter_rows(min_row=2):
        total += 1
        row_vals = [cell.value for cell in fila]

        def get_by_key(key):
            idx = found.get(key)
            if idx is None or idx >= len(row_vals):
                return None
            return row_vals[idx]

        def limpiar(valor):
            return '' if valor in (None, '-', '–') else str(valor).strip()

        codigo = limpiar(get_by_key('codigo'))
        descripcion = limpiar(get_by_key('descripcion'))
        partida = limpiar(get_by_key('partida'))


        chapter_from_code = _detect_chapter_from_code(codigo)

        combined = get_by_key('ace22_chi_prot')
        chi_val = get_by_key('ace22_chi')
        prot_val = get_by_key('ace22_prot')
        if combined:
            ace22_val = limpiar(combined)
        else:
            chi = limpiar(chi_val)
            prot = limpiar(prot_val)
            if chi and prot:
                ace22_val = f"{chi}; {prot}"
            elif chi:
                ace22_val = chi
            elif prot:
                ace22_val = prot
            else:
                ace22_val = ''

        datos = {
            'capitulo': chapter_from_code or '',
            'partida': partida,
            'codigo': codigo,
            'descripcion': descripcion,
            'gravamen': limpiar(get_by_key('gravamen')),
            'ice_iehd': limpiar(get_by_key('ice_iehd')),
            'unidad_medida': limpiar(get_by_key('unidad_medida')),
            'despacho_frontera': limpiar(get_by_key('despacho_frontera')),
            'tipo_documento': limpiar(get_by_key('tipo_documento')),
            'entidad_emite': limpiar(get_by_key('entidad_emite')),
            'disp_legal': limpiar(get_by_key('disp_legal')),
            'can_ace36_ace47_ven': limpiar(get_by_key('can_ace36_ace47_ven')),
            'ace66_mexico': limpiar(get_by_key('ace66_mexico')),
            'subpartida': limpiar(get_by_key('subpartida')),
            'ace22': ace22_val
        }


        if not codigo or not descripcion:
            errors.append(f"Fila {total+1}: datos faltantes (codigo/descripcion)")
            continue

        if codigo in codes_in_file:
            errors.append(f"Fila {total+1}: codigo duplicado en archivo ({codigo})")
            continue
        codes_in_file.add(codigo)

        if not chapter_from_code:
            errors.append(f"Fila {total+1}: codigo sin prefijo válido (debe empezar con 2 dígitos)")
            continue


        chapter_key = chapter_from_code
        if chapter_key not in rows_by_chapter:
            rows_by_chapter[chapter_key] = []
        rows_by_chapter[chapter_key].append(datos)

    if errors:
        return {'total': total, 'imported': 0, 'created': 0, 'updated': 0, 'deleted': 0, 'omitted': len(errors), 'errors': errors, 'log': None}


    if not sync_catalog:
        applied_created = 0
        applied_updated = 0
        applied_omitted = 0
        save_errors = []
        
        for chapter_data in rows_by_chapter.values():
            for datos2 in chapter_data:
                codigo2 = datos2.get('codigo')
                try:
                    existing = PartidaArancelaria.objects.filter(codigo=codigo2).first()
                    if existing:
                        if update_existing:
                            for f in ['capitulo','partida','descripcion','gravamen','ice_iehd','unidad_medida','despacho_frontera','tipo_documento','entidad_emite','disp_legal','can_ace36_ace47_ven','ace66_mexico','subpartida']:
                                val = datos2.get(f)
                                if val:
                                    setattr(existing, f, val)
                            if datos2.get('ace22'):
                                existing.ace22_chi_prot = datos2.get('ace22')
                            existing.save()
                            applied_updated += 1
                        else:
                            applied_omitted += 1
                        continue
                    PartidaArancelaria.objects.create(
                        capitulo=datos2.get('capitulo') or '',
                        partida=datos2.get('partida') or '',
                        codigo=codigo2,
                        descripcion=datos2.get('descripcion') or '',
                        gravamen=datos2.get('gravamen') or '',
                        ice_iehd=datos2.get('ice_iehd') or '',
                        unidad_medida=datos2.get('unidad_medida') or '',
                        despacho_frontera=datos2.get('despacho_frontera') or '',
                        tipo_documento=datos2.get('tipo_documento') or '',
                        entidad_emite=datos2.get('entidad_emite') or '',
                        disp_legal=datos2.get('disp_legal') or '',
                        can_ace36_ace47_ven=datos2.get('can_ace36_ace47_ven') or '',
                        ace22_chi_prot=datos2.get('ace22') or '',
                        ace66_mexico=datos2.get('ace66_mexico') or '',
                        subpartida=datos2.get('subpartida') or ''
                    )
                    applied_created += 1
                except Exception as e:
                    save_errors.append(f"Error fila {codigo2}: {e}")
        
        try:
            nombre = nombre_archivo or getattr(source, 'name', None) or 'uploaded'
            log = ImportLog.objects.create(usuario=usuario, nombre_archivo=nombre, total_filas=total, importadas=(applied_created+applied_updated), omitidas=applied_omitted, errores='\n'.join(save_errors) if save_errors else '')
        except Exception:
            log = None
        return {'total': total, 'imported': (applied_created+applied_updated), 'created': applied_created, 'updated': applied_updated, 'deleted': 0, 'omitted': applied_omitted, 'errors': save_errors, 'log': log}


    total_created = 0
    total_updated = 0
    total_deleted = 0
    all_errors = []

    for chapter_target, rows_data in rows_by_chapter.items():
        excel_codes = set([d['codigo'] for d in rows_data])
        

        existing_qs = PartidaArancelaria.objects.filter(codigo__startswith=str(chapter_target))
        existing_map = {p.codigo: p for p in existing_qs}


        for datos2 in rows_data:
            codigo2 = datos2.get('codigo')
            try:
                existing = existing_map.get(codigo2)
                if existing:

                    changed = False
                    mapping = {
                        'capitulo': 'capitulo',
                        'partida': 'partida',
                        'descripcion': 'descripcion',
                        'gravamen': 'gravamen',
                        'ice_iehd': 'ice_iehd',
                        'unidad_medida': 'unidad_medida',
                        'despacho_frontera': 'despacho_frontera',
                        'tipo_documento': 'tipo_documento',
                        'entidad_emite': 'entidad_emite',
                        'disp_legal': 'disp_legal',
                        'can_ace36_ace47_ven': 'can_ace36_ace47_ven',
                        'ace22': 'ace22_chi_prot',
                        'ace66_mexico': 'ace66_mexico',
                        'subpartida': 'subpartida'
                    }
                    for src_field, dest_field in mapping.items():
                        new_val = datos2.get(src_field) or ''
                        old_val = getattr(existing, dest_field) or ''
                        if str(new_val).strip() != str(old_val).strip():
                            changed = True
                            setattr(existing, dest_field, new_val)
                    if changed:
                        existing.save()
                        total_updated += 1
                else:

                    cap_to_set = datos2.get('capitulo') or chapter_target
                    PartidaArancelaria.objects.create(
                        capitulo=cap_to_set or '',
                        partida=datos2.get('partida') or '',
                        codigo=codigo2,
                        descripcion=datos2.get('descripcion') or '',
                        gravamen=datos2.get('gravamen') or '',
                        ice_iehd=datos2.get('ice_iehd') or '',
                        unidad_medida=datos2.get('unidad_medida') or '',
                        despacho_frontera=datos2.get('despacho_frontera') or '',
                        tipo_documento=datos2.get('tipo_documento') or '',
                        entidad_emite=datos2.get('entidad_emite') or '',
                        disp_legal=datos2.get('disp_legal') or '',
                        can_ace36_ace47_ven=datos2.get('can_ace36_ace47_ven') or '',
                        ace22_chi_prot=datos2.get('ace22') or '',
                        ace66_mexico=datos2.get('ace66_mexico') or '',
                        subpartida=datos2.get('subpartida') or ''
                    )
                    total_created += 1
            except Exception as e:
                all_errors.append(f"Error capítulo {chapter_target}, código {codigo2}: {e}")


        try:
            qs_to_delete = PartidaArancelaria.objects.filter(codigo__startswith=str(chapter_target)).exclude(codigo__in=list(excel_codes))
            deleted_count = qs_to_delete.count()
            if deleted_count:
                qs_to_delete.delete()
                total_deleted += deleted_count
        except Exception as e:
            all_errors.append(f"Error eliminación capítulo {chapter_target}: {e}")


    try:
        nombre = nombre_archivo or getattr(source, 'name', None) or 'uploaded'
        summary_errors = '\n'.join(all_errors) if all_errors else ''
        log = ImportLog.objects.create(
            usuario=usuario,
            nombre_archivo=nombre,
            total_filas=total,
            importadas=(total_created + total_updated),
            omitidas=0,
            errores=(summary_errors + (f"\nSincronización: creadas={total_created}, actualizadas={total_updated}, eliminadas={total_deleted}" if (total_created or total_updated or total_deleted) else '')).strip()
        )
    except Exception:
        log = None

    return {'total': total, 'imported': (total_created + total_updated), 'created': total_created, 'updated': total_updated, 'deleted': total_deleted, 'omitted': 0, 'errors': all_errors, 'log': log}
