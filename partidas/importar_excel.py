import openpyxl
import re
import unicodedata
import tempfile
import os
from .models import PartidaArancelaria, ImportLog


def _normalize(s: str) -> str:
    s = (s or '')
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r'[^0-9a-z]+', ' ', s)
    s = ' '.join(s.split())
    return s


def _split_normalize_ace22(full_ace22: str):
    if not full_ace22:
        return '', ''
    s = str(full_ace22).strip()
    sep_candidates = [';', '|', '/', ',', '\\n']
    split_parts = None
    for sep in sep_candidates:
        if sep in s:
            split_parts = [p.strip() for p in s.split(sep) if p.strip()]
            break
    if not split_parts:
        if '  ' in s:
            split_parts = [p.strip() for p in s.split('  ') if p.strip()]
    if not split_parts:
        parts = [s]
    else:
        parts = split_parts

    bad_set = {'n', 'y', 's', 'si', 'no', '0', '1', 'yes', 'a'}
    def clean(val):
        if not val:
            return ''
        v = val.strip()
        if len(v) <= 1 and v.lower() in bad_set:
            return ''
        if v.lower() in bad_set:
            return ''
        return v

    chi = clean(parts[0]) if len(parts) >= 1 else ''
    prot = clean(parts[1]) if len(parts) >= 2 else ''
    return chi, prot


def _detect_headers(sheet):
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
    encabezados_raw = [cell.value if cell.value is not None else '' for cell in header_cells]
    encabezados = [_normalize(h) for h in encabezados_raw]

    variantes = {
        'capitulo': ['capitulo', 'capítulo'],
        'partida': ['partida'],
        'subpartida': ['subpartida'],
        'codigo': ['codigo', 'código'],
        'descripcion': ['descripcion', 'descripcion de la mercancia', 'descripcion de la mercadería', 'descripcion de la mercancía'],
        'gravamen': ['gravamen', 'impuesto'],
        'ice_iehd': ['ice iehd', 'ice', 'iehd', 'ice_iehd'],
        'unidad_medida': ['unidad de medida', 'unidad_medida', 'unidad'],
        'despacho_frontera': ['despacho en frontera', 'despacho frontera', 'despacho_frontera'],
        'tipo_documento': ['tipo de documento', 'tipo_documento'],
        'entidad_emite': ['entidad que emite', 'entidad_emite', 'entidad'],
        'disp_legal': ['disposicion legal', 'disposición legal', 'disp_legal', 'disposicion'],
        'can_ace36_ace47_ven': ['can ace36 ace47 ven', 'ace36 ace47', 'can_ace36_ace47_ven'],
        'ace22_chi_prot': ['ace22 chi prot', 'ace22_chi_prot', 'ace22'],
        'ace22_chi': ['ace22 chi', 'ace22_chi'],
        'ace22_prot': ['ace22 prot', 'ace22_prot'],
        'ace66_mexico': ['ace66 mexico', 'ace66_mexico', 'ace66']
    }

    header_index = {h: idx for idx, h in enumerate(encabezados)}

    def find_index_for(variants):
        for v in variants:
            nv = _normalize(v)
            if nv in header_index:
                return header_index[nv]
        for v in variants:
            nv = _normalize(v)
            for h, idx in header_index.items():
                if nv in h or h in nv:
                    return idx
        return None

    found = {}
    required_min = ['capitulo', 'partida', 'codigo', 'descripcion']
    for campo in required_min:
        idx = find_index_for(variantes.get(campo, [campo]))
        if idx is None:
            raise Exception(f"El archivo no contiene la columna requerida: '{campo}'")
        found[campo] = idx

    ace22_idx = find_index_for(variantes.get('ace22_chi_prot'))
    ace22_chi_idx = find_index_for(variantes.get('ace22_chi'))
    ace22_prot_idx = find_index_for(variantes.get('ace22_prot'))
    if ace22_idx is None and ace22_chi_idx is None and ace22_prot_idx is None:
        # no es crítico, solo advertiremos en preview
        pass

    optional_fields = ['gravamen','ice_iehd','unidad_medida','despacho_frontera','tipo_documento','entidad_emite','disp_legal','can_ace36_ace47_ven','ace66_mexico','subpartida']
    for of in optional_fields:
        idx = find_index_for(variantes.get(of, [of]))
        if idx is not None:
            found[of] = idx

    if ace22_idx is not None:
        found['ace22_chi_prot'] = ace22_idx
    if ace22_chi_idx is not None:
        found['ace22_chi'] = ace22_chi_idx
    if ace22_prot_idx is not None:
        found['ace22_prot'] = ace22_prot_idx

    return found


def preview_import(source):
    """Devuelve una estructura con filas detectadas y errores por fila para mostrar en preview.
    `source` puede ser un path a archivo o un file-like.
    """
    if isinstance(source, str):
        wb = openpyxl.load_workbook(source)
    else:
        wb = openpyxl.load_workbook(source)
    sheet = wb.active
    found = _detect_headers(sheet)

    rows = []
    seen_codes = set()
    line_no = 1
    for fila in sheet.iter_rows(min_row=2):
        line_no += 1
        row_vals = [cell.value for cell in fila]

        def get_by_key(key):
            idx = found.get(key)
            if idx is None or idx >= len(row_vals):
                return None
            return row_vals[idx]

        def limpiar(valor):
            return '' if valor in (None, '-', '–') else str(valor).strip()

        codigo = limpiar(get_by_key('codigo'))
        descripcion = limpiar(get_by_key('descripcion'))
        capitulo = limpiar(get_by_key('capitulo'))
        partida = limpiar(get_by_key('partida'))

        ace22_val = ''
        combined = get_by_key('ace22_chi_prot')
        chi_val = get_by_key('ace22_chi')
        prot_val = get_by_key('ace22_prot')
        if combined:
            ace22_val = limpiar(combined)
        else:
            chi = limpiar(chi_val)
            prot = limpiar(prot_val)
            if chi and prot:
                ace22_val = f"{chi}; {prot}"
            elif chi:
                ace22_val = chi
            elif prot:
                ace22_val = prot

        errors = []
        if not codigo:
            errors.append('codigo vacío')
        if not descripcion:
            errors.append('descripcion vacía')
        if not capitulo:
            errors.append('capitulo vacío')
        if not partida:
            errors.append('partida vacío')

        if codigo:
            if codigo in seen_codes:
                errors.append('codigo duplicado en archivo')
            seen_codes.add(codigo)
            if PartidaArancelaria.objects.filter(codigo=codigo).exists():
                errors.append('codigo ya existe en la base de datos')

        rows.append({'line': line_no, 'codigo': codigo, 'descripcion': descripcion, 'capitulo': capitulo, 'partida': partida, 'ace22': ace22_val, 'errors': errors})

    total = len(rows)
    errors_count = sum(1 for r in rows if r['errors'])
    return {'rows': rows, 'total': total, 'errors_count': errors_count}


def process_import(source, usuario=None, update_existing=False, nombre_archivo=None):
    """Procesa la importación real y crea un ImportLog.
    Retorna un dict resumen con counts y mensajes de error.
    """
    if isinstance(source, str):
        wb = openpyxl.load_workbook(source)
    else:
        wb = openpyxl.load_workbook(source)
    sheet = wb.active
    found = _detect_headers(sheet)

    total = 0
    imported = 0
    omitted = 0
    errors = []

    for fila in sheet.iter_rows(min_row=2):
        total += 1
        row_vals = [cell.value for cell in fila]

        def get_by_key(key):
            idx = found.get(key)
            if idx is None or idx >= len(row_vals):
                return None
            return row_vals[idx]

        def limpiar(valor):
            return '' if valor in (None, '-', '–') else str(valor).strip()

        datos = {
            'capitulo': limpiar(get_by_key('capitulo')),
            'partida': limpiar(get_by_key('partida')),
            'codigo': limpiar(get_by_key('codigo')),
            'descripcion': limpiar(get_by_key('descripcion')),
            'gravamen': limpiar(get_by_key('gravamen')),
            'ice_iehd': limpiar(get_by_key('ice_iehd')),
            'unidad_medida': limpiar(get_by_key('unidad_medida')),
            'despacho_frontera': limpiar(get_by_key('despacho_frontera')),
            'tipo_documento': limpiar(get_by_key('tipo_documento')),
            'entidad_emite': limpiar(get_by_key('entidad_emite')),
            'disp_legal': limpiar(get_by_key('disp_legal')),
            'can_ace36_ace47_ven': limpiar(get_by_key('can_ace36_ace47_ven')),
            'ace66_mexico': limpiar(get_by_key('ace66_mexico')),
            'subpartida': limpiar(get_by_key('subpartida'))
        }

        combined = get_by_key('ace22_chi_prot')
        chi_val = get_by_key('ace22_chi')
        prot_val = get_by_key('ace22_prot')
        if combined:
            ace22_val = limpiar(combined)
        else:
            chi = limpiar(chi_val)
            prot = limpiar(prot_val)
            if chi and prot:
                ace22_val = f"{chi}; {prot}"
            elif chi:
                ace22_val = chi
            elif prot:
                ace22_val = prot
            else:
                ace22_val = ''

        codigo = datos.get('codigo')
        if not codigo or not datos.get('descripcion'):
            omitted += 1
            errors.append(f"Fila {total+1}: datos faltantes (codigo/descripcion)")
            continue

        try:
            existing = PartidaArancelaria.objects.filter(codigo=codigo).first()
            if existing:
                if update_existing:
                    existing.capitulo = datos.get('capitulo') or existing.capitulo
                    existing.partida = datos.get('partida') or existing.partida
                    existing.descripcion = datos.get('descripcion') or existing.descripcion
                    existing.gravamen = datos.get('gravamen') or existing.gravamen
                    existing.ice_iehd = datos.get('ice_iehd') or existing.ice_iehd
                    existing.unidad_medida = datos.get('unidad_medida') or existing.unidad_medida
                    existing.despacho_frontera = datos.get('despacho_frontera') or existing.despacho_frontera
                    existing.tipo_documento = datos.get('tipo_documento') or existing.tipo_documento
                    existing.entidad_emite = datos.get('entidad_emite') or existing.entidad_emite
                    existing.disp_legal = datos.get('disp_legal') or existing.disp_legal
                    existing.can_ace36_ace47_ven = datos.get('can_ace36_ace47_ven') or existing.can_ace36_ace47_ven
                    existing.ace22_chi_prot = ace22_val or existing.ace22_chi_prot
                    existing.ace66_mexico = datos.get('ace66_mexico') or existing.ace66_mexico
                    existing.save()
                    imported += 1
                else:
                    omitted += 1
                continue

            PartidaArancelaria.objects.create(
                capitulo=datos.get('capitulo') or '',
                partida=datos.get('partida') or '',
                codigo=codigo,
                descripcion=datos.get('descripcion') or '',
                gravamen=datos.get('gravamen') or '',
                ice_iehd=datos.get('ice_iehd') or '',
                unidad_medida=datos.get('unidad_medida') or '',
                despacho_frontera=datos.get('despacho_frontera') or '',
                tipo_documento=datos.get('tipo_documento') or '',
                entidad_emite=datos.get('entidad_emite') or '',
                disp_legal=datos.get('disp_legal') or '',
                can_ace36_ace47_ven=datos.get('can_ace36_ace47_ven') or '',
                ace22_chi_prot=ace22_val,
                ace66_mexico=datos.get('ace66_mexico') or ''
            )
            imported += 1
        except Exception as e:
            omitted += 1
            errors.append(f"Fila {total+1}: excepción al guardar -> {e}")

    # crear registro de ImportLog
    try:
        nombre = nombre_archivo or getattr(source, 'name', None) or 'uploaded'
        log = ImportLog.objects.create(
            usuario=usuario,
            nombre_archivo=nombre,
            total_filas=total,
            importadas=imported,
            omitidas=omitted,
            errores='\n'.join(errors) if errors else ''
        )
    except Exception:
        log = None

    return {'total': total, 'imported': imported, 'omitted': omitted, 'errors': errors, 'log': log}

