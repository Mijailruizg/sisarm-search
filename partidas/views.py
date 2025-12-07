from django.shortcuts import render, get_object_or_404, redirect
from .models import PartidaArancelaria, Busqueda, Manual, LicenciaTemporal, Rol, PartidaReferencia, HistoriaActividad, Usuario
from .forms import CargarExcelForm, PartidaForm, RegistroUsuarioForm, UsuarioAdminForm
from .importar_excel import preview_import, process_import
import tempfile
import os
from .decorators import rol_requerido
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.contrib import messages
from django.utils.timezone import now
from django.contrib.auth import login
from django.db.models import Q
from datetime import timedelta, datetime
from django.utils.safestring import mark_safe
import re
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.core.mail import send_mail
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
import csv
from .decorators import rol_requerido
from .models import ExportLog
from .models import ClickLog

from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

# Cliente de chat: las importaciones siguientes causan errores, se comentan
# El sistema usa _generate_local_reply() que se define más abajo
get_chat_response = None
stream_chat_response = None
DEFAULT_RESPONSE = 'Lo siento, el asistente no está disponible en este momento.'


def _split_normalize_ace22(full_ace22: str):
    """Divide el campo ACE22 combinado en (chi, prot) y normaliza valores "vacíos".
    Considera como vacíos valores muy cortos o marcadores como 'N', 'Y', 'SI', 'NO', '0', '1'.
    Devuelve tupla (chi, prot) donde cada valor es cadena o '' si no aplicable.
    """
    if not full_ace22:
        return '', ''
    s = str(full_ace22).strip()
    sep_candidates = [';', '|', '/', ',', '\\n']
    split_parts = None
    for sep in sep_candidates:
        if sep in s:
            split_parts = [p.strip() for p in s.split(sep) if p.strip()]
            break
    if not split_parts:
        if '  ' in s:
            split_parts = [p.strip() for p in s.split('  ') if p.strip()]
    if not split_parts:
        parts = [s]
    else:
        parts = split_parts

    # normalizar y filtrar valores no informativos
    # incluir variantes de indicadores no informativos (letras sueltas usadas en algunos Excel/ingresos)
    bad_set = {'n', 'y', 's', 'si', 'no', '0', '1', 'yes', 'no', 'a'}
    def clean(val):
        if not val:
            return ''
        v = val.strip()
        if len(v) <= 1 and v.lower() in bad_set:
            return ''
        # también filtrar sí/no abreviados de 2 letras
        if v.lower() in bad_set:
            return ''
        return v

    chi = clean(parts[0]) if len(parts) >= 1 else ''
    prot = clean(parts[1]) if len(parts) >= 2 else ''
    return chi, prot

def registro(request):
    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            try:
                usuario = form.save(commit=False)
                try:
                    rol_usuario = Rol.objects.get(nombre__iexact='usuario')
                    usuario.rol = rol_usuario
                except Rol.DoesNotExist:
                    messages.error(request, "El rol 'usuario' no existe. Contacta al administrador.")
                    return redirect('registro')

                usuario.save()

                if not LicenciaTemporal.objects.filter(usuario=usuario).exists():
                    # usar date (sin hora) para evitar discrepancias de zona/hora al calcular días
                    from datetime import date as _date
                    fecha_inicio = _date.today()
                    # Fecha fin calculada para dar 7 días de vigencia INCLUSIVOS (hoy + 6 días)
                    fecha_fin = fecha_inicio + timedelta(days=6)
                    LicenciaTemporal.objects.create(
                        usuario=usuario,
                        fecha_inicio=fecha_inicio,
                        fecha_fin=fecha_fin,
                        estado=True
                    )
                    # enviar email de bienvenida con fecha de vencimiento (si el usuario tiene email)
                    try:
                        if usuario.email:
                            subject = 'Bienvenido a SISARM - licencia temporal activada'
                            body = f"Hola {usuario.username},\n\nTu licencia temporal ha sido activada hasta el {fecha_fin}. Tendrás acceso completo durante 7 días.\n\nSaludos,\nEquipo SISARM"
                            from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', None)
                            # si no hay DEFAULT_FROM_EMAIL, send_mail acepta None pero algunos backends lo requieren; intentar con '' si None
                            send_mail(subject, body, from_email or '', [usuario.email], fail_silently=True)
                    except Exception:
                        # no interrumpir el registro si falla el envío de correo
                        pass

                login(request, usuario)
                return redirect('inicio')
            except Exception as e:
                messages.error(request, f"Ocurrió un error al registrar: {e}")
                return redirect('registro')
    else:
        form = RegistroUsuarioForm()

    return render(request, 'registration/registro.html', {'form': form})


@login_required
def inicio(request):
    from datetime import date
    from .models import LicenciaTemporal
    licencia = LicenciaTemporal.objects.filter(usuario=request.user, estado=True).order_by('-fecha_fin').first()
    dias_licencia = None
    fecha_fin_iso = None
    if licencia:
        hoy = date.today()
        raw_days = (licencia.fecha_fin - hoy).days
        # Mostrar días inclusivos: remaining = raw_days + 1 cuando raw_days >= 0
        dias_licencia = (raw_days + 1) if raw_days >= 0 else 0
        # pasar fecha de fin completa en formato ISO para el contador JS
        fecha_fin_iso = licencia.fecha_fin.isoformat()
        # notificar al usuario si la licencia vence en pocos días (3 días o menos)
        try:
            # mostrar mensajes en la UI (advertir si quedan 1..3 días)
            if raw_days >= 0 and dias_licencia <= 3:
                messages.warning(request, f"Tu licencia vence en {dias_licencia} día{'' if dias_licencia==1 else 's'} (hasta {licencia.fecha_fin}). Por favor renueva para evitar interrupciones.")
                # además, enviar un email una sola vez si no fue enviado aún
                try:
                    if not getattr(licencia, 'notified_pre_expiry', False) and request.user.email:
                        subject = 'Aviso: licencia SISARM próxima a vencer'
                        body = f"Hola {request.user.username},\n\nTu licencia de SISARM vence el {licencia.fecha_fin} (faltan {dias_licencia} día{'' if dias_licencia==1 else 's'}). Por favor renueva para evitar interrupciones.\n\nSaludos,\nEquipo SISARM"
                        from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', None)
                        send_mail(subject, body, from_email or '', [request.user.email], fail_silently=True)
                        licencia.notified_pre_expiry = True
                        licencia.save()
                except Exception:
                    pass
            elif raw_days < 0:
                # licencia ya expirada
                messages.error(request, f"Tu licencia expiró el {licencia.fecha_fin}. Algunas funciones pueden estar limitadas.")
        except Exception:
            pass
    return render(request, 'partidas/inicio.html', {
        'dias_licencia': dias_licencia
        , 'fecha_fin_iso': fecha_fin_iso
    })

@login_required
def buscar_partidas(request):
    partidas = PartidaArancelaria.objects.all()
    termino = request.GET.get('termino', '').strip()
    capitulo = request.GET.get('capitulo', '').strip()
    gravamen = request.GET.get('gravamen', '').strip()
    tipo_documento = request.GET.get('tipo_documento', '').strip()
    entidad_emite = request.GET.get('entidad_emite', '').strip()
    disp_legal = request.GET.get('disp_legal', '').strip()

    # Aplicar filtros (sin requerir término)
    if termino:
        partidas = partidas.filter(Q(codigo__icontains=termino) | Q(descripcion__icontains=termino))

    if capitulo:
        partidas = partidas.filter(capitulo__icontains=capitulo)
    if gravamen:
        partidas = partidas.filter(gravamen__icontains=gravamen)
    if tipo_documento:
        partidas = partidas.filter(tipo_documento__iexact=tipo_documento)
    if entidad_emite:
        partidas = partidas.filter(entidad_emite__icontains=entidad_emite)
    if disp_legal:
        partidas = partidas.filter(disp_legal__iexact=disp_legal)

    # Registrar búsqueda solo si hay término de búsqueda
    if request.user.is_authenticated and termino:
        # preparar resumen de resultados: cantidad y hasta 5 códigos como ejemplo
        try:
            total_hits = partidas.count()
            ejemplos = list(partidas.values_list('codigo', flat=True)[:5])
            ejemplos_txt = ', '.join(ejemplos) if ejemplos else 'sin resultados'
            resumen = f"{total_hits} resultados; Ej: {ejemplos_txt}"
        except Exception:
            # en caso de cualquier problema con el queryset, guardar una nota genérica
            resumen = 'No disponible'

        Busqueda.objects.create(
            usuario=request.user,
            termino_buscado=termino,
            tipo_busqueda="Texto o Código",
            fecha=now(),
            resultados=resumen
        )
        # Actualizar contadores: diario y total acumulado por capítulo
        from .models import SearchStatisticDaily, SearchStatisticTotal, SearchStatisticProductTotal
        from datetime import date as _date
        hoy = _date.today()
        matches_for_stats = PartidaArancelaria.objects.filter(Q(codigo__icontains=termino) | Q(descripcion__icontains=termino))
        chapters_seen = set()
        for m in matches_for_stats:
            chap = m.capitulo or 'Sin capítulo'
            if chap in chapters_seen:
                continue
            chapters_seen.add(chap)
            # diario
            d_obj, _ = SearchStatisticDaily.objects.get_or_create(capitulo=chap, fecha=hoy, defaults={'count': 0})
            d_obj.count = d_obj.count + 1
            d_obj.save()
            # total acumulado
            t_obj, _ = SearchStatisticTotal.objects.get_or_create(capitulo=chap, defaults={'total': 0})
            t_obj.total = t_obj.total + 1
            t_obj.save()
            # total por producto
            # usar el código de la partida como identificador único
            prod_key = m.codigo or f"{m.id}"
            p_obj, _ = SearchStatisticProductTotal.objects.get_or_create(
                codigo=prod_key,
                defaults={'descripcion': m.descripcion, 'capitulo': chap, 'total': 0}
            )
            p_obj.total = p_obj.total + 1
            # mantener la descripción y capítulo actualizados si cambien
            p_obj.descripcion = m.descripcion
            p_obj.capitulo = chap
            p_obj.save()

    capitulo_relacionado = partidas.first().capitulo if partidas.exists() else None
    relacionadas = PartidaArancelaria.objects.filter(
        capitulo=capitulo_relacionado
    ).exclude(id__in=partidas)[:15] if capitulo_relacionado else []

    similares = []
    if termino:
        similares = PartidaArancelaria.objects.filter(
            descripcion__icontains=termino
        ).exclude(id__in=partidas.values_list('id', flat=True))[:15]

    if termino:
        patron = re.compile(re.escape(termino), re.IGNORECASE)
        for p in partidas:
            p.descripcion_resaltada = mark_safe(
                patron.sub(
                    lambda m: f'<mark style="background-color:#00ff55; color:#000; padding:0.2em 0.3em; border-radius:5px;">{m.group(0)}</mark>',
                    p.descripcion
                )
            )
        for s in similares:
            s.descripcion_resaltada = mark_safe(
                patron.sub(
                    lambda m: f'<mark style="background-color:#00ff55; color:#000; padding:0.2em 0.3em; border-radius:5px;">{m.group(0)}</mark>',
                    s.descripcion
                )
            )
    else:
        for p in partidas:
            p.descripcion_resaltada = p.descripcion
        for s in similares:
            s.descripcion_resaltada = s.descripcion

    # Separar ace22_chi_prot en dos atributos para mostrar en la tabla de resultados (ACE22 CHI / ACE22 PROT)
    try:
        for p in partidas:
            full_ace22 = (p.ace22_chi_prot or '').strip()
            ace22_chi, ace22_prot = _split_normalize_ace22(full_ace22)
            # asignar atributos dinámicos para la plantilla
            setattr(p, 'ace22_chi', ace22_chi)
            setattr(p, 'ace22_prot', ace22_prot)
    except Exception:
        # no bloquear la búsqueda si falla la separación; dejar los atributos vacíos
        pass

    # obtener capítulos distintos y preparar etiquetas limpias para mostrar en el filtro
    raw_capitulos = list(PartidaArancelaria.objects.values_list('capitulo', flat=True).distinct())
    def clean_capitulo_label(c):
        if not c:
            return None
        s = str(c).strip()
        if not s:
            return None
        # eliminar prefijos como 'Capitulo 1:', 'Capítulo 1 -', '1 -', '01.' etc.
        s_clean = re.sub(r'^(capitulo|capículo|capítulo)\s*\d+\s*[:\-\)]?\s*', '', s, flags=re.IGNORECASE)
        s_clean = re.sub(r'^\d+\s*[:\.\-\)]\s*', '', s_clean)
        s_clean = s_clean.strip()
        return s_clean or None

    capitulos_disponibles = []
    seen = set()
    for c in raw_capitulos:
        # ignorar valores vacíos o genéricos
        if c is None:
            continue
        orig = str(c).strip()
        if not orig or orig.lower() in ('sin datos', 'n/a'):
            continue
        if orig in seen:
            continue
        seen.add(orig)
        label = clean_capitulo_label(orig)
        # intentar extraer número de capítulo (primer entero que aparezca)
        num_match = re.search(r"(\d+)", orig)
        numero = int(num_match.group(1)) if num_match else None
        # si no pudimos limpiar, usar el original
        if not label:
            label = orig
        capitulos_disponibles.append((orig, label, numero))
    
    # Obtener gravamenes_disponibles y eliminar duplicados
    gravamenes_raw = PartidaArancelaria.objects.values_list('gravamen', flat=True).distinct()
    gravamenes_set = set()
    gravamenes_disponibles = []
    for g in gravamenes_raw:
        if g and g not in gravamenes_set:
            gravamenes_set.add(g)
            gravamenes_disponibles.append(g)
    
    # Obtener tipos_disponibles y eliminar duplicados (SQLite puede devolver duplicados con .distinct())
    tipos_raw = list(PartidaArancelaria.objects.values_list('tipo_documento', flat=True).distinct())
    tipos_set = set()
    tipos_disponibles = []
    for t in tipos_raw:
        if t and t not in tipos_set:
            tipos_set.add(t)
            tipos_disponibles.append(t)
    
    # Obtener entidades_disponibles y eliminar duplicados
    entidades_raw = list(PartidaArancelaria.objects.values_list('entidad_emite', flat=True).distinct())
    entidades_set = set()
    entidades_disponibles = []
    for e in entidades_raw:
        if e and e not in entidades_set:
            entidades_set.add(e)
            entidades_disponibles.append(e)
    
    # Obtener disp_legal_disponibles y eliminar duplicados
    disp_legal_raw = list(PartidaArancelaria.objects.values_list('disp_legal', flat=True).distinct())
    disp_legal_set = set()
    disp_legal_disponibles = []
    for d in disp_legal_raw:
        if d and d not in disp_legal_set:
            disp_legal_set.add(d)
            disp_legal_disponibles.append(d)

    return render(request, 'partidas/buscar.html', {
        'resultados': partidas,
        'termino': termino,
        'relacionadas': relacionadas,
        'similares': similares,
        'capitulo': capitulo_relacionado,
        'filtros': {
            'capitulo': capitulo,
            'gravamen': gravamen,
            'tipo_documento': tipo_documento,
            'entidad_emite': entidad_emite,
            'disp_legal': disp_legal
        },
        'capitulos': capitulos_disponibles,
        'gravamenes': gravamenes_disponibles,
        'tipos_doc': tipos_disponibles,
        'entidades': entidades_disponibles,
        'disp_legales': disp_legal_disponibles
    })


def api_autocomplete(request):
    """Endpoint simple de autocompletado usado por la UI.
    Devuelve JSON con lista de objetos {codigo, descripcion}.
    Estrategia:
      1) Buscar por código que empiece por la query (prioridad)
      2) Luego buscar por código que contenga o descripción que contenga
      3) Si quedan espacios, intentar búsqueda por tokens en la descripción
    """
    q = (request.GET.get('q') or '').strip()
    if not q:
        return JsonResponse({'results': []})

    max_results = 15
    results = []

    try:
        # Tomar un conjunto amplio de candidatos y puntuar en Python para ordenar por relevancia.
        # Esto funciona con SQLite (sin pg_trgm).
        tokens = [t.strip().lower() for t in q.split() if t.strip()]
        q_lower = q.lower()

        # traer candidatos que contengan la query en código o descripción (limitar a 200 para rendimiento)
        candidates = PartidaArancelaria.objects.filter(
            Q(codigo__icontains=q) | Q(descripcion__icontains=q)
        ).order_by('codigo')[:200]

        scored = []
        seen = set()
        for p in candidates:
            if p.id in seen:
                continue
            seen.add(p.id)
            score = 0
            codigo = (p.codigo or '').lower()
            descripcion = (p.descripcion or '').lower()

            # Puntos por tipo de coincidencia (mayor es mejor)
            if codigo == q_lower:
                score += 300
            if codigo.startswith(q_lower):
                score += 150
            if q_lower in codigo and not codigo.startswith(q_lower):
                score += 80

            # coincidencia en descripción
            if q_lower in descripcion:
                score += 60

            # tokens: más tokens coincidentes => mejor puntuación
            token_matches = 0
            for t in tokens:
                if t in descripcion:
                    token_matches += 1
            score += token_matches * 12

            # bonus por longitud de código (códigos más cortos pueden ser más generales)
            try:
                if codigo:
                    score += max(0, 10 - len(codigo))
            except Exception:
                pass

            scored.append((score, p))

        # Ordenar por score descendente y luego por código ascendente
        scored.sort(key=lambda x: (-x[0], x[1].codigo or ''))

        for score, p in scored[:max_results]:
            results.append({'codigo': p.codigo, 'descripcion': p.descripcion or ''})

    except Exception as e:
        try:
            print('api_autocomplete error:', e)
        except Exception:
            pass

    return JsonResponse({'results': results})


# Webhook endpoint para Dialogflow fulfillment (recibe requests desde Dialogflow)
@csrf_exempt
def dialogflow_webhook(request):
    try:
        import json
        req = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return JsonResponse({'fulfillmentText': 'Error parsing request'}, status=400)

    # Extraer intent name
    intent = req.get('queryResult', {}).get('intent', {}).get('displayName', '')
    session = req.get('session', '')  # formato: projects/.../agent/sessions/<sessionId>

    # Intent específico: estado de licencia
    try:
        if 'license' in intent.lower() or 'licencia' in intent.lower():
            # intentar extraer user id desde session si usamos formato 'user-<id>' en el session_id
            session_id = session.split('/')[-1] if session else ''
            user_id = None
            if session_id.startswith('user-'):
                try:
                    user_id = int(session_id.split('user-')[-1])
                except Exception:
                    user_id = None
            # intentar buscar licencia en DB si user_id disponible
            if user_id:
                try:
                    licencia = LicenciaTemporal.objects.filter(usuario__id=user_id, estado=True).order_by('-fecha_fin').first()
                    if licencia:
                        from datetime import date
                        hoy = date.today()
                        dias = max(0, (licencia.fecha_fin - hoy).days + 1)
                        text = f'Tu licencia vence el {licencia.fecha_fin}. Quedan {dias} día{"s" if dias!=1 else ""}.'
                        return JsonResponse({'fulfillmentText': text})
                except Exception:
                    pass
            # si no hay user o licencia, devolver instrucción genérica
            return JsonResponse({'fulfillmentText': 'Para ver el estado de tu licencia, inicia sesión en el sistema o solicita soporte desde la sección Soporte.'})

        # Intent de búsqueda de partida: si Dialogflow envía parámetro 'codigo' lo buscamos
        if 'partida' in intent.lower() or 'buscar' in intent.lower():
            params = req.get('queryResult', {}).get('parameters', {}) or {}
            codigo = params.get('codigo') or params.get('number') or params.get('any')
            if codigo:
                try:
                    p = PartidaArancelaria.objects.filter(codigo__icontains=str(codigo)).first()
                    if p:
                        text = f"Partida {p.codigo}: {p.descripcion[:300]}"
                        return JsonResponse({'fulfillmentText': text})
                except Exception:
                    pass

        # Fallback: delegar al generador local para respuestas contextuales
        try:
            query_text = req.get('queryResult', {}).get('queryText', '')
            reply, _, _ = _generate_local_reply(query_text, request=None)
            return JsonResponse({'fulfillmentText': reply})
        except Exception:
            return JsonResponse({'fulfillmentText': 'Lo siento, no pude procesar tu solicitud ahora.'})

    except Exception as e:
        return JsonResponse({'fulfillmentText': f'Error interno: {e}'}, status=500)

@login_required
def detalle_partida(request, partida_id):
    try:
        partida = get_object_or_404(PartidaArancelaria, id=partida_id)

        # intentar extraer metadatos simples de la referencia legal (fecha, número, nota corta)
        referencia_text = partida.referencia_legal or partida.disp_legal or ''
        referencia_fecha = None
        referencia_numero = None
        referencia_nota = None
        try:
            # buscar año (ej. 2020) como indicación de fecha
            m_year = re.search(r"(19|20)\d{2}", referencia_text)
            if m_year:
                referencia_fecha = m_year.group(0)
            # buscar número de resolución (Nº, No., Número)
            m_num = re.search(r"N[\u00BA\u00AA]?\s*[:\.]?\s*(\d+)|No\.?\s*(\d+)|Número\s*(\d+)", referencia_text, re.IGNORECASE)
            if m_num:
                for g in m_num.groups():
                    if g:
                        referencia_numero = g
                        break
            referencia_nota = referencia_text.strip()[:250]
        except Exception:
            referencia_text = referencia_text

        context = {
            'partida': partida,
            'referencia_text': referencia_text,
            'referencia_fecha': referencia_fecha,
            'referencia_numero': referencia_numero,
            'referencia_nota': referencia_nota,
        }
        # Separar ace22_chi_prot en dos valores para mostrar en la plantilla
        full_ace22 = (partida.ace22_chi_prot or '').strip()
        ace22_chi, ace22_prot = _split_normalize_ace22(full_ace22)
        context['ace22_chi'] = ace22_chi
        context['ace22_prot'] = ace22_prot
        # adjuntar referencias/documentos asociados (si los hay)
        referencias = PartidaReferencia.objects.filter(partida=partida).order_by('-creado_en')
        context['referencias'] = referencias

        # Exportar PDF cuando se pide via ?export=pdf
        if request.GET.get('export') == 'pdf':
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.pdfgen import canvas
                from io import BytesIO

                buffer = BytesIO()
                c = canvas.Canvas(buffer, pagesize=letter)
                width, height = letter

                title = f"Partida {partida.codigo} - SISARM Search"
                c.setFont('Helvetica-Bold', 14)
                c.drawString(40, height - 60, title)

                c.setFont('Helvetica', 11)
                y = height - 90
                lines = [
                    ("Código:", partida.codigo or ""),
                    ("Partida:", partida.partida or ""),
                    ("Descripción:", (partida.descripcion or "").replace('\n', ' ')),
                    ("Gravamen:", partida.gravamen or ""),
                    ("ICE / IEHD:", partida.ice_iehd or ""),
                    ("Unidad de Medida:", partida.unidad_medida or ""),
                    ("Despacho Frontera:", partida.despacho_frontera or ""),
                    ("Tipo Documento:", partida.tipo_documento or ""),
                    ("Entidad que Emite:", partida.entidad_emite or ""),
                    ("Disposición Legal (resumen):", (partida.disp_legal or '')[:400]),
                    ("Capítulo:", partida.capitulo or ""),
                ]

                for label, val in lines:
                    text = f"{label} {val}"
                    max_chars = 95
                    while text:
                        part = text[:max_chars]
                        c.drawString(40, y, part)
                        y -= 14
                        text = text[max_chars:]
                        if y < 80:
                            c.showPage()
                            c.setFont('Helvetica', 11)
                            y = height - 60

                if referencias:
                    if y < 100:
                        c.showPage()
                        c.setFont('Helvetica', 11)
                        y = height - 60
                    c.setFont('Helvetica-Bold', 12)
                    c.drawString(40, y, "Referencias / Documentos asociados:")
                    y -= 18
                    c.setFont('Helvetica', 10)
                    for ref in referencias:
                        ref_line = f"- {ref.titulo or ref.numero_resolucion or 'Referencia'} ({ref.fecha_norma or ''})"
                        text = ref_line
                        max_chars = 95
                        while text:
                            part = text[:max_chars]
                            c.drawString(44, y, part)
                            y -= 12
                            text = text[max_chars:]
                            if y < 60:
                                c.showPage()
                                c.setFont('Helvetica', 10)
                                y = height - 60

                c.showPage()
                c.save()
                pdf = buffer.getvalue()
                buffer.close()

                response = HttpResponse(content_type='application/pdf')
                filename = f"partida_{partida.codigo or partida.id}.pdf"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                response.write(pdf)
                return response
            except Exception:
                try:
                    print('PDF export error while generating PDF for detalle_partida')
                except Exception:
                    pass

        return render(request, 'partidas/detalle_partida.html', context)
    except Exception as e:
        # devolver un mensaje claro a la plantilla si ocurre un error
        try:
            msg = f"Error al cargar los datos de la partida: {str(e)}"
        except Exception:
            msg = 'Error al cargar los datos de la partida.'
        return render(request, 'partidas/detalle_partida.html', {'load_error': msg})

@login_required
def historial_buscador(request):
    # Manejar eliminación vía POST
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'delete_selected':
            ids = request.POST.getlist('selected')
            if ids:
                Busqueda.objects.filter(usuario=request.user, id__in=ids).delete()
        elif action == 'delete_all':
            Busqueda.objects.filter(usuario=request.user).delete()
        return redirect('historial_buscador')

    # Obtener últimas 20 búsquedas del usuario y buscar partidas relacionadas para cada término
    historial_qs = Busqueda.objects.filter(usuario=request.user).order_by('-fecha')[:20]
    historial = []
    for b in historial_qs:
        matches = PartidaArancelaria.objects.filter(
            Q(codigo__icontains=b.termino_buscado) | Q(descripcion__icontains=b.termino_buscado)
        ).order_by('codigo')[:15]
        # Normalizar ACE22 para cada partida en matches (para mostrar columnas separadas en plantilla)
        try:
            for p in matches:
                chi, prot = _split_normalize_ace22(p.ace22_chi_prot or '')
                setattr(p, 'ace22_chi', chi)
                setattr(p, 'ace22_prot', prot)
        except Exception:
            pass
        historial.append({'item': b, 'matches': matches})
    return render(request, 'partidas/historial.html', {'historial': historial})

@login_required
@rol_requerido('Administrador')
def importar_excel(request):
    # Flujo: 1) Subir archivo -> preview 2) Confirmar -> procesar
    if request.method == 'POST':
        # Confirmación final
        if request.POST.get('confirm') == '1':
            tmp_path = request.session.get('import_tmp_path')
            tmp_name = request.session.get('import_tmp_name')
            if not tmp_path or not os.path.exists(tmp_path):
                messages.error(request, 'No se encontró el archivo temporal de importación. Vuelve a subir el archivo.')
                return redirect('importar_excel')

            update_existing = bool(request.POST.get('update_existing'))
            result = process_import(tmp_path, usuario=request.user, update_existing=update_existing, nombre_archivo=tmp_name)

            # limpiar archivo temporal y sesión
            try:
                os.remove(tmp_path)
            except Exception:
                pass
            request.session.pop('import_tmp_path', None)
            request.session.pop('import_tmp_name', None)

            # Mensajes con el resumen
            messages.success(request, f"Importación completada: {result.get('imported',0)} importadas, {result.get('omitted',0)} omitidas de {result.get('total',0)} filas.")
            if result.get('errors'):
                messages.error(request, 'Algunas filas tuvieron errores. Revisa el registro de importación en el admin.')
            return redirect('panel_partidas')

        # Primera subida: generar preview
        form = CargarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            # guardar temporalmente en disco para poder reabrir en confirmación
            tf = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            try:
                for chunk in archivo.chunks():
                    tf.write(chunk)
            finally:
                tf.close()

            request.session['import_tmp_path'] = tf.name
            request.session['import_tmp_name'] = getattr(archivo, 'name', 'uploaded.xlsx')

            preview = preview_import(tf.name)
            return render(request, 'partidas/import_preview.html', {'preview': preview, 'archivo_nombre': request.session['import_tmp_name']})
    else:
        form = CargarExcelForm()
    return render(request, 'partidas/importar_excel.html', {'form': form})

@rol_requerido('Administrador')
def panel_partidas(request):
    partidas = PartidaArancelaria.objects.all()
    return render(request, 'partidas/panel_partidas.html', {'partidas': partidas})


@rol_requerido('Administrador')
def crear_partida(request):
    if request.method == 'POST':
        form = PartidaForm(request.POST)
        if form.is_valid():
            partida = form.save()
            try:
                HistoriaActividad.objects.create(usuario=request.user, accion=f"admin: crear partida {partida.codigo} (id={partida.pk})")
            except Exception:
                pass
            messages.success(request, 'Partida creada correctamente.')
            return redirect('panel_partidas')
    else:
        form = PartidaForm()
    return render(request, 'partidas/partida_form.html', {'form': form, 'titulo': 'Crear Partida'})


@rol_requerido('Administrador')
def editar_partida(request, partida_id):
    partida = get_object_or_404(PartidaArancelaria, pk=partida_id)
    if request.method == 'POST':
        form = PartidaForm(request.POST, instance=partida)
        if form.is_valid():
            part = form.save()
            try:
                HistoriaActividad.objects.create(usuario=request.user, accion=f"admin: editar partida {part.codigo} (id={part.pk})")
            except Exception:
                pass
            messages.success(request, 'Partida actualizada correctamente.')
            return redirect('panel_partidas')
    else:
        form = PartidaForm(instance=partida)
    return render(request, 'partidas/partida_form.html', {'form': form, 'titulo': 'Editar Partida'})


@rol_requerido('Administrador')
def eliminar_partida(request, partida_id):
    partida = get_object_or_404(PartidaArancelaria, pk=partida_id)
    if request.method == 'POST':
        codigo = partida.codigo
        partida.delete()
        try:
            HistoriaActividad.objects.create(usuario=request.user, accion=f"admin: eliminar partida {codigo} (id={partida_id})")
        except Exception:
            pass
        messages.success(request, 'Partida eliminada correctamente.')
        return redirect('panel_partidas')
    return render(request, 'partidas/partida_confirm_delete.html', {'partida': partida})


@rol_requerido('Administrador')
def admin_usuarios(request):
    """Lista de usuarios para administradores con acciones básicas."""
    qs = Usuario.objects.all().order_by('username')
    return render(request, 'partidas/admin_usuarios_list.html', {'usuarios': qs})


@rol_requerido('Administrador')
def crear_usuario(request):
    if request.method == 'POST':
        form = UsuarioAdminForm(request.POST)
        if form.is_valid():
            user = form.save()
            try:
                HistoriaActividad.objects.create(usuario=request.user, accion=f"admin: crear usuario {user.username} (id={user.pk})")
            except Exception:
                pass
            messages.success(request, 'Usuario creado correctamente.')
            return redirect('admin_usuarios')
    else:
        form = UsuarioAdminForm()
    return render(request, 'partidas/usuario_form.html', {'form': form, 'titulo': 'Crear usuario'})


@rol_requerido('Administrador')
def editar_usuario(request, usuario_id):
    user = get_object_or_404(Usuario, pk=usuario_id)
    if request.method == 'POST':
        form = UsuarioAdminForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            try:
                HistoriaActividad.objects.create(usuario=request.user, accion=f"admin: editar usuario {user.username} (id={user.pk})")
            except Exception:
                pass
            messages.success(request, 'Usuario actualizado correctamente.')
            return redirect('admin_usuarios')
    else:
        form = UsuarioAdminForm(instance=user)
    return render(request, 'partidas/usuario_form.html', {'form': form, 'titulo': 'Editar usuario'})


@rol_requerido('Administrador')
def toggle_usuario_activo(request, usuario_id):
    user = get_object_or_404(Usuario, pk=usuario_id)
    user.is_active = not user.is_active
    user.save()
    try:
        HistoriaActividad.objects.create(usuario=request.user, accion=f"admin: {'activar' if user.is_active else 'desactivar'} usuario {user.username} (id={user.pk})")
    except Exception:
        pass
    messages.success(request, f"Usuario {'activado' if user.is_active else 'desactivado'}.")
    return redirect('admin_usuarios')

@login_required
def ver_manuales(request):
    # Soporta búsqueda por palabra clave (q) sobre tipo y descripción
    q = (request.GET.get('q') or '').strip()
    manuales_qs = Manual.objects.all().order_by('-updated_at', 'tipo')
    if q:
        manuales_qs = manuales_qs.filter(Q(tipo__icontains=q) | Q(descripcion__icontains=q))

    # Agrupar por tipo (módulo) para navegación más intuitiva
    manuals_by_type = {}
    manual_usuario = Manual.objects.filter(tipo__iexact='manual de usuario').first()
    for m in manuales_qs:
        # excluir la 'Guía del Buscador' del listado principal (se muestra por separado)
        if (m.tipo or '').strip().lower() == 'guía del buscador' or (m.tipo or '').strip().lower() == 'guia del buscador':
            continue
        key = m.tipo or 'General'
        manuals_by_type.setdefault(key, []).append(m)

    return render(request, 'partidas/manuales.html', {
        'manuales': manuales_qs,
        'manuales_por_tipo': manuals_by_type,
        'q': q
    })


@login_required
def guia_buscador_html(request):
    """Vista ligera que muestra la Guía del Buscador en HTML (más accesible que el PDF)."""
    return render(request, 'partidas/manuales_guia_buscador.html', {})


@login_required
def descargar_manual_usuario(request):
    """Genera y devuelve un PDF del Manual de Usuario (versión imprimible).
    Construye un PDF sencillo a partir de texto estructurado usando ReportLab y lo devuelve como attachment.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.units import mm
        from io import BytesIO
    except Exception:
        return HttpResponse('Generación de PDF no disponible: falta la librería reportlab.', status=500)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    h1 = styles['Heading1']
    h2 = styles['Heading2']
    normal = styles['BodyText']
    normal.alignment = TA_LEFT

    story = []
    story.append(Paragraph('Manual de Usuario — SISARM Search', h1))
    story.append(Spacer(1, 6))

    # Secciones breves (coinciden con la versión HTML)
    story.append(Paragraph('Sección A — Inicio y configuración de la cuenta', h2))
    story.append(Paragraph('1. Crear cuenta: ve a Crear nueva cuenta y completa nombre, apellido, correo, usuario y contraseña.', normal))
    story.append(Paragraph('2. Inicio de sesión: usa tus credenciales en Iniciar sesión. Si olvidas la contraseña usa recuperación en el formulario de login.', normal))
    story.append(Paragraph('3. Perfil: desde la esquina superior, revisa tu correo y datos. Mantén el correo actualizado para notificaciones.', normal))
    story.append(Spacer(1, 6))

    story.append(Paragraph('Sección B — Buscar partidas (flujo básico)', h2))
    story.append(Paragraph('1. En el menú principal selecciona Buscar Partidas.', normal))
    story.append(Paragraph('2. Escribe el código completo (ej. 010121) o una palabra clave de la descripción.', normal))
    story.append(Paragraph('3. Pulsa Buscar. Los resultados se muestran con columnas esenciales (Código, Descripción, Gravamen).', normal))
    story.append(Paragraph('4. Si necesitas precisión, haz clic sobre el resultado para abrir la Ficha de detalle.', normal))
    story.append(Spacer(1, 6))

    story.append(Paragraph('Sección C — Uso de filtros (mejorar resultados)', h2))
    story.append(Paragraph('• Capítulo: selecciona el capítulo para limitar familias de productos (ej. Capítulo 1).', normal))
    story.append(Paragraph('• Gravamen / Impuestos: filtra por texto que aparezca en el campo gravamen.', normal))
    story.append(Paragraph('• Entidad que emite: filtra por la entidad (Ministerio, Instituto, etc.).', normal))
    story.append(Spacer(1, 6))

    story.append(Paragraph('Sección D — Ficha de detalle de una partida', h2))
    story.append(Paragraph('• En la ficha verás: código completo, descripción amplia y campos adicionales (gravamen, ICE/IEHD, documentos requeridos y entidad emisora).', normal))
    story.append(Paragraph("• ACE22: el sistema muestra ACE22 dividido en CHI y PROT cuando existe; si no hay dato mostrará 'N/A'.", normal))
    story.append(Paragraph('• Documentos: revisa la lista de referencias y usa «Solicitar ayuda» si falta un documento.', normal))
    story.append(Spacer(1, 6))

    story.append(Paragraph('Sección E — Historial y exportaciones', h2))
    story.append(Paragraph('• Historial: guarda búsquedas recientes con fecha; útil para repetir consultas.', normal))
    story.append(Paragraph('• Exportar a Excel: los administradores pueden exportar resultados (archivo con columnas claramente separadas y formato listo para análisis).', normal))
    story.append(Spacer(1, 6))

    story.append(Paragraph('Sección F — Licencias y acceso', h2))
    story.append(Paragraph('• Las cuentas de prueba reciben licencia temporal de 7 días (incluyentes).', normal))
    story.append(Paragraph('• Revisa el panel de inicio para ver días restantes; el contador es inclusivo y se actualiza cada minuto.', normal))
    story.append(Spacer(1, 6))

    story.append(Paragraph('Sección G — Soporte y ayuda', h2))
    story.append(Paragraph('• Asistente: usa Chat de Ayuda para preguntas rápidas (licencia, filtros, soporte).', normal))
    story.append(Paragraph('• Soporte: abre el formulario de Soporte desde el menú y adjunta capturas o ejemplos.', normal))

    # Construir PDF
    try:
        doc.build(story)
    except Exception as e:
        return HttpResponse(f'Error al generar PDF: {e}', status=500)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="manual_usuario_sisarm.pdf"'
    return response


def lista_aranceles(request):
    """Índice de capítulos disponibles."""
    capitulos_qs = PartidaArancelaria.objects.values_list('capitulo', flat=True).distinct().order_by('capitulo')
    capitulos = []
    for c in capitulos_qs:
        capitulos.append({'capitulo': c, 'count': PartidaArancelaria.objects.filter(capitulo=c).count()})
    return render(request, 'partidas/lista_aranceles.html', {'capitulos': capitulos})


@login_required
@rol_requerido('Administrador')
def exportar_partidas_excel(request):
    """Exporta partidas a Excel (.xlsx) con las columnas mostradas en la imagen del usuario.
    Acepta filtros GET similares a la búsqueda: termino, capitulo, gravamen, tipo_documento, entidad_emite
    """
    qs = PartidaArancelaria.objects.all().order_by('codigo')

    termino = request.GET.get('termino', '').strip()
    capitulo = request.GET.get('capitulo', '').strip()
    gravamen = request.GET.get('gravamen', '').strip()
    tipo_documento = request.GET.get('tipo_documento', '').strip()
    entidad_emite = request.GET.get('entidad_emite', '').strip()

    if termino:
        qs = qs.filter(Q(codigo__icontains=termino) | Q(descripcion__icontains=termino))
    if capitulo:
        qs = qs.filter(capitulo__icontains=capitulo)
    if gravamen:
        qs = qs.filter(gravamen__icontains=gravamen)
    if tipo_documento:
        qs = qs.filter(tipo_documento__iexact=tipo_documento)
    if entidad_emite:
        qs = qs.filter(entidad_emite__icontains=entidad_emite)

    # Registrar la exportación
    filtros = {'termino': termino, 'capitulo': capitulo, 'gravamen': gravamen, 'tipo_documento': tipo_documento, 'entidad_emite': entidad_emite}
    try:
        ExportLog.objects.create(usuario=request.user, accion='export_partidas_xlsx', filtros=str(filtros))
    except Exception:
        pass

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Partidas'

    headers = [
        'capitulo', 'partida', 'codigo', 'descripcion de la mercancia', 'gravamen',
        'ice - iehd', 'unidad de medida', 'despacho en frontera', 'tipo de documento',
        'entidad que emite', 'disposicion legal', 'can ace36 ace47 ven', 'ace22_chi_prot', 'ace66_mexico'
    ]

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Escribir cabecera con estilo
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Ajustar: exportar ACE22 en dos columnas separadas (CHI / PROT)
    headers = [
        'capitulo', 'partida', 'codigo', 'descripcion de la mercancia', 'gravamen',
        'ice - iehd', 'unidad de medida', 'despacho en frontera', 'tipo de documento',
        'entidad que emite', 'disposicion legal', 'can ace36 ace47 ven', 'ace22_chi', 'ace22_prot', 'ace66_mexico'
    ]

    # reescribir cabecera (sobrescribir previamente escrita si es necesario)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Rellenar filas
    row = 2
    for p in qs.iterator():
        # intentar separar el campo combinado ace22_chi_prot en dos columnas
        full_ace22 = (p.ace22_chi_prot or '').strip()
        ace22_chi, ace22_prot = _split_normalize_ace22(full_ace22)

        # Si están vacíos después de normalizar, dejar 'N/A' para mayor claridad en el Excel
        ace22_chi_val = ace22_chi if ace22_chi else 'N/A'
        ace22_prot_val = ace22_prot if ace22_prot else 'N/A'

        # Valores en el orden pedido: Capítulo primero, luego Partida, Código, ... y ACE22 combinado
        values = [
            p.capitulo or '',
            p.partida or '',
            p.codigo or '',
            (p.descripcion or '').replace('\r', ' ').replace('\n', ' '),
            p.gravamen or '',
            p.ice_iehd or '',
            p.unidad_medida or '',
            p.despacho_frontera or '',
            p.tipo_documento or '',
            p.entidad_emite or '',
            p.disp_legal or '',
            p.can_ace36_ace47_ven or '',
            ace22_chi_val,
            ace22_prot_val,
            p.ace66_mexico or ''
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            # ajustar alineación para columnas largas
            if col_idx in (4, 11):
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                cell.alignment = Alignment(vertical='top')
            cell.border = border
        row += 1

    # Ajustar anchos de columna básicos
    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        # descripción (ahora columna 4) y disposición legal (columna 11) son largas
        if i == 4 or i == 11:
            ws.column_dimensions[col_letter].width = 80
        elif i == 3:
            ws.column_dimensions[col_letter].width = 20
        else:
            ws.column_dimensions[col_letter].width = 16

    # Guardar en un BytesIO y retornar response
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = 'partidas_export.xlsx'
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def aranceles_por_capitulo(request, capitulo):
    qs = PartidaArancelaria.objects.filter(capitulo=capitulo).order_by('codigo')
    return render(request, 'partidas/aranceles_por_capitulo.html', {'capitulo': capitulo, 'aranceles': qs})


# Custom LoginView: after a successful login, redirect to licencia_expirada
# when the user's latest license is missing, expired or inactive.
class CustomLoginView(LoginView):
    def form_valid(self, form):
        # login the user first
        response = super().form_valid(form)
        try:
            user = self.request.user
            from datetime import date
            licencia = None
            try:
                licencia = user.licenciatemporal_set.order_by('-fecha_fin').first()
            except Exception:
                licencia = None
            if (not licencia) or (licencia and licencia.fecha_fin < date.today()) or (licencia and not getattr(licencia, 'estado', True)):
                return redirect('licencia_expirada')
        except Exception:
            # on any unexpected error, fall back to default behaviour
            return response
        return response

def chat_asistente(request):
    # conservar el texto original para mostrarlo en la plantilla
    original_mensaje = request.GET.get('mensaje', '').strip()
    mensaje = original_mensaje.lower()
    # timestamp para la respuesta (se mostrará en la plantilla)
    from django.utils import timezone
    respuesta_tiempo = timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')

    # delegar la lógica de generación local a una función reutilizable
    respuesta, sugerencias, action = _generate_local_reply(mensaje, request)

    context = {
        # mostrar el mensaje original enviado por el usuario en la plantilla
        "mensaje": original_mensaje,
        "respuesta": respuesta,
        "sugerencias": sugerencias
    }
    # Si el frontend pide abrir soporte vía query param, preparar la acción
    # (permite enlaces desde emails que abran el chat y muestren la burbuja/modal de Soporte)
    open_support_q = request.GET.get('open_support') or request.GET.get('action')
    if open_support_q and str(open_support_q).lower() in ('1', 'true', 'yes', 'open_support'):
        # pasar parametros prefill si vienen en GET (email, prefill_subject, prefill_message)
        action = {'open_support': '/soporte/', 'action_text': 'Abrir la página de Soporte.'}
        # solicitar apertura automática cuando el enlace viene con open_support
        action['auto_open'] = True
        email_prefill = request.GET.get('email')
        subj_prefill = request.GET.get('prefill_subject')
        msg_prefill = request.GET.get('prefill_message')
        if email_prefill:
            action['email'] = email_prefill
        if subj_prefill:
            action['prefill_subject'] = subj_prefill
        if msg_prefill:
            action['prefill_message'] = msg_prefill
    # acción opcional que la plantilla puede ejecutar en el cliente (abrir soporte, whatsapp, etc.)
    import json
    # enriquecer la action con un texto legible para la UI (action_text)
    action_payload = None
    if action and isinstance(action, dict):
        action_payload = dict(action)
        # si el backend no incluyó una descripción, generarla aquí
        if 'action_text' not in action_payload:
            if 'open_support' in action_payload:
                action_payload['action_text'] = 'Abrir la página de Soporte.'
            elif 'open_whatsapp' in action_payload:
                action_payload['action_text'] = 'Abrir WhatsApp para contactar al soporte.'
            else:
                action_payload['action_text'] = 'Abrir recurso sugerido.'

    context["action_json"] = json.dumps(action_payload) if action_payload else 'null'
    # incluir timestamp para mostrar cuándo se generó la respuesta
    context["respuesta_tiempo"] = respuesta_tiempo
    return render(request, 'partidas/chat.html', context)

def licencia_expirada(request):
    return render(request, 'partidas/licencia_expirada.html')


def solicitar_renovacion(request):
    """Vista pública para solicitar renovación o contactar soporte desde la página de licencia expirada.
    GET: muestra un formulario para indicar correo y mensaje.
    POST: crea `SolicitudSoporte`, intenta enviar el correo y redirige a `licencia_expirada` con mensaje.
    """
    from .models import SolicitudSoporte

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        # Priorizar subject_override si el JS creó ese campo para asunto personalizado
        subject = (request.POST.get('subject_override') or request.POST.get('subject') or '').strip() or f"Solicitud de renovación de licencia"
        message = request.POST.get('message', '').strip()

        # si el usuario está autenticado y no proporcionó email, usar el email del usuario
        if not email and request.user.is_authenticated:
            email = getattr(request.user, 'email', '') or ''

        if not email:
            messages.error(request, 'Debes indicar un correo de contacto para que podamos responderte.')
            return redirect('solicitar_renovacion')

        if not message:
            # mensaje por defecto si no se proporciona
            message = f"Solicitud de renovación iniciada desde la página de Licencia Expirada. Usuario: {request.user.username if request.user.is_authenticated else 'Anónimo'}"

        nombre_val = request.POST.get('nombre', '').strip() if request.method == 'POST' else ''
        if not nombre_val and request.user.is_authenticated:
            nombre_val = getattr(request.user, 'get_full_name', lambda: '')() or getattr(request.user, 'username', '')

        try:
            solicitud = SolicitudSoporte.objects.create(
                usuario=request.user if request.user.is_authenticated else None,
                nombre=nombre_val or None,
                correo=email or 'sin-correo@local',
                asunto=subject,
                mensaje=message,
                estado='pending'
            )
        except Exception:
            solicitud = None

        # Intento de envío
        try:
            support_to = getattr(settings, 'SUPPORT_EMAIL', 'soporte@sisarm.com')
            send_mail(subject, message, email or getattr(settings, 'DEFAULT_FROM_EMAIL', ''), [support_to], fail_silently=False)
            # No marcar automáticamente como 'sent' — mantener como 'pending' para que el admin procese y confirme.
            messages.success(request, 'Solicitud registrada. El equipo de soporte la procesará.')
        except Exception as e:
            # Registrar el error en el registro sin cambiar el estado (permanece 'pending')
            if solicitud:
                solicitud.error_message = str(e)
                solicitud.save()
            messages.error(request, f'No se pudo enviar la notificación por correo: {e}. Su petición fue registrada y el equipo la revisará.')

        return redirect('licencia_expirada')

    # GET: mostrar formulario simple para recopilar email y mensaje
    initial_email = request.user.email if request.user.is_authenticated else ''
    initial_name = ''
    if request.user.is_authenticated:
        initial_name = getattr(request.user, 'get_full_name', lambda: '')() or getattr(request.user, 'username', '')
    return render(request, 'partidas/solicitar_renovacion.html', {'initial_email': initial_email, 'initial_name': initial_name})

from django.shortcuts import render

@login_required
def soporte(request):
    # Proveer valores iniciales para autofill en la plantilla
    initial_email = request.user.email if request.user.is_authenticated else ''
    initial_name = ''
    if request.user.is_authenticated:
        initial_name = getattr(request.user, 'get_full_name', lambda: '')() or getattr(request.user, 'username', '')
    return render(request, 'partidas/soporte.html', {'initial_email': initial_email, 'initial_name': initial_name})


@login_required
def soporte_submit(request):
    """Recibe POST desde el formulario de Soporte y lo envía por email al equipo de soporte.
    Guarda también una entrada en HistoriaActividad.
    """
    from django.shortcuts import redirect
    from django.contrib import messages
    # Priorizar subject_override para asuntos personalizados enviados desde el formulario
    subject = (request.POST.get('subject_override') or request.POST.get('subject') or '').strip()
    message = request.POST.get('message', '').strip()
    email_from = request.POST.get('email') or (request.user.email if request.user.is_authenticated else '')
    if not email_from:
        messages.error(request, 'Debes indicar un correo de contacto.')
        return redirect('soporte')
    if not message:
        messages.error(request, 'El mensaje no puede estar vacío.')
        return redirect('soporte')

    # Registrar la solicitud en la base de datos inmediatamente (fecha/hora automática)
    try:
        from .models import SolicitudSoporte
        nombre_val = request.POST.get('nombre', '').strip()
        if not nombre_val and request.user.is_authenticated:
            nombre_val = getattr(request.user, 'get_full_name', lambda: '')() or getattr(request.user, 'username', '')
        solicitud = SolicitudSoporte.objects.create(
            usuario=request.user if request.user.is_authenticated else None,
            nombre=nombre_val or None,
            correo=email_from,
            asunto=subject or 'Sin asunto',
            mensaje=message,
            estado='pending'
        )
    except Exception:
        solicitud = None

    # Construir cuerpo con datos de contacto
    full_body = f"Solicitud de soporte desde {email_from}\n\n{message}"
    try:
        # enviar al equipo de soporte; ajustar en settings si es necesario
        support_to = getattr(settings, 'SUPPORT_EMAIL', 'soporte@sisarm.com')
        try:
            sent_count = send_mail(subject or 'Solicitud de soporte desde la aplicación', full_body, email_from, [support_to], fail_silently=False)
            # registrar actividad
            try:
                HistoriaActividad.objects.create(usuario=request.user if request.user.is_authenticated else None, accion=f"soporte_submit: {subject[:200]}")
            except Exception:
                pass
            # marcar como enviado si el envío fue exitoso
            try:
                if solicitud:
                    solicitud.estado = 'sent'
                    solicitud.error_message = ''
                    solicitud.save()
            except Exception:
                pass
            messages.success(request, 'Solicitud registrada y enviada al equipo de soporte. Gracias.')
        except Exception as e:
            # si falla el envío por email, marcar la solicitud con error y guardar mensaje
            try:
                if solicitud:
                    solicitud.estado = 'error'
                    solicitud.error_message = str(e)
                    solicitud.save()
            except Exception:
                pass
            messages.error(request, f'Error al enviar la solicitud: {str(e)}')
    except Exception as e:
        # si falla el envío por email, marcar la solicitud con error y guardar mensaje
        try:
            if solicitud:
                solicitud.error_message = str(e)
                solicitud.save()
        except Exception:
            pass
        messages.error(request, f'Error al enviar la solicitud: {str(e)}')

    return redirect('soporte')


# función local que genera la respuesta del asistente sin depender de Dialogflow/OpenAI
def _generate_local_reply(mensaje: str, request):
    """Genera una respuesta corta usando reglas locales. Devuelve (respuesta, sugerencias, action).
    action es un diccionario opcional con instrucciones que el cliente puede ejecutar, por ejemplo:
      {'open_support': '/soporte/'} o {'open_whatsapp': 'https://wa.me/...'}
    """
    respuesta = ""
    sugerencias = []
    action = None
    # normalizar: sin acentos, en minúsculas y sin espacios extremos
    import unicodedata
    raw = (mensaje or '')
    m = raw.strip().lower()
    # eliminar acentos para mejorar matching
    try:
        m = ''.join(c for c in unicodedata.normalize('NFKD', m) if not unicodedata.combining(c))
    except Exception:
        pass
    # correcciones simples de typos comunes
    corrections = {
        'ola': 'hola',
        'ta bien': 'esta bien',
        'esta bien': 'esta bien',
        'gracais': 'gracias'
    }
    if m in corrections:
        m = corrections[m]

    # Si el usuario responde afirmativamente y hay una acción sugerida en sesión, ejecutarla
    try:
        affirmatives = {'si', 'sí', 's', 'ok', 'vale', 'dale', 'claro'}
        if m in affirmatives:
            last_action = None
            try:
                last_action = request.session.pop('chat_last_action')
            except Exception:
                last_action = None
            if last_action:
                respuesta = 'Abriendo la página solicitada...'
                sugerencias = []
                action = last_action
                return respuesta, sugerencias, action
    except Exception:
        # si falla cualquier manejo de sesión, continuar normalmente
        pass

    # intent recognition: preguntas frecuentes y variantes
    # Mejores patrones para detectar WhatsApp, filtros y consultas sobre impuestos
    if 'whatsapp' in m or 'atienden por whatsapp' in m or 'whats' in m:
        # Respuesta mejorada: indicar si hay atención por WhatsApp y ofrecer abrir la conversación
        # Leer número y texto desde settings si están definidos para personalizar
        try:
            wa_number = getattr(settings, 'SUPPORT_WHATSAPP_NUMBER', '59177682918')
            wa_text = getattr(settings, 'SUPPORT_WHATSAPP_TEXT', 'Hola, necesito ayuda con SISARM Search.')
        except Exception:
            wa_number = '59177682918'
            wa_text = 'Hola, necesito ayuda con SISARM Search.'
        try:
            from urllib.parse import quote
            wa_link = f'https://wa.me/{wa_number}?text={quote(wa_text)}'
        except Exception:
            wa_link = f'https://wa.me/{wa_number}'
        respuesta = (
            "Actualmente el canal principal de soporte es el formulario del sistema y el correo <strong>soporte@sisarm.com</strong>. "
            "Si prefieres WhatsApp, puedes abrir la conversación con nuestro número haciendo clic en el enlace o te lo puedo abrir ahora."
            f" <br><br><a href=\"{wa_link}\" target=\"_blank\" class=\"btn btn-success\">Abrir WhatsApp</a>"
        )
        sugerencias = ["Contactar soporte", "Ver manuales"]
        # ofrecer que el cliente abra WhatsApp automáticamente
        action = {'open_whatsapp': wa_link, 'action_text': f'Abrir WhatsApp ({wa_number})'}
        return respuesta, sugerencias, action

    if 'donde' in m or 'dónde' in m or 'donde se ve' in m or 'dónde se ve' in m or 'donde veo' in m or 'dónde veo' in m:
        # preguntas sobre dónde ver impuestos o gravámenes
        if 'gravamen' in m or 'impuesto' in m or 'ice' in m or 'iehd' in m:
            respuesta = (
                "En la ficha de cada partida (detalle de partida) verás los campos relacionados con gravámenes e impuestos: 'gravamen' muestra el valor o texto del impuesto aplicable; si corresponde, también aparece la etiqueta ICE/IEHD en el campo 'ice_iehd'."
                " Busca la partida desde 'Buscar Partidas' y haz clic en el resultado para ver esos campos en su detalle."
            )
            sugerencias = ["Buscar partida", "Ver detalle de partida"]
            return respuesta, sugerencias

    # preguntas sobre filtros
    if 'filtro' in m or 'filtrar' in m or 'filtros' in m or 'capítulo' in m and 'fil' in m:
        respuesta = (
            "Para acotar resultados utiliza los filtros disponibles en la pantalla de búsqueda:\n"
            "• Capítulo: selecciona la familia de productos (ej. 01, 02…).\n"
            "• Gravamen: filtra por el texto o valor del impuesto mostrado en la partida.\n"
            "• Entidad: filtra por la entidad emisora (Ministerio, Instituto, etc.).\n"
            "Aplica uno o varios filtros simultáneamente para reducir la lista de resultados."
        )
        sugerencias = ["Buscar partida", "Filtrar por capítulo"]
        return respuesta, sugerencias

    if not m:
        respuesta = "¡Hola! Soy el Asistente de SISARM. Puedo ayudarte a buscar partidas, ver manuales o revisar tu licencia. ¿Qué necesitas hacer?"
        sugerencias = ["Buscar partida", "Ver manuales", "Mi licencia", "Contactar soporte"]

    elif m in ["hola", "buenas", "buenos días", "buenas tardes", "buenas noches"]:
        respuesta = "¡Hola! ¿En qué puedo ayudarte hoy? Puedo guiarte a buscar partidas, revisar manuales o ver el estado de tu licencia."
        sugerencias = ["Buscar partida", "Ver manuales", "Mi licencia"]

    elif m in ["gracias", "muchas gracias", "ok", "entiendo"]:
        respuesta = "Con gusto. Si necesitas otra cosa, dime: buscar una partida, ver manuales o contactar soporte."
        sugerencias = ["Buscar partida", "Mi licencia", "Contactar soporte"]

    elif m in ["adiós", "chau", "hasta luego"]:
        respuesta = "¡Hasta luego! Cuando necesites ayuda vuelve a escribir."
        sugerencias = []

    elif "qué es una subpartida" in m or m == "subpartida":
        respuesta = (
            "Una subpartida es una subdivisión más detallada dentro de una partida arancelaria. Permite clasificar productos con mayor precisión y definir requisitos específicos."
        )
        sugerencias = ["¿Dónde veo los documentos requeridos?", "Buscar partida"]

    elif "dónde veo los documentos" in m or "ver documentos requeridos" in m:
        respuesta = (
            "Para ver los documentos y requisitos, busca la partida y haz clic en el resultado para ver su detalle."
            " En el detalle verás: tipos de documento, entidad emisora y la disposición legal relacionada."
        )
        sugerencias = ["Buscar partida", "¿Qué entidad emite el permiso?"]

    elif "qué entidad emite el permiso" in m:
        respuesta = (
            "La entidad emisora (por ejemplo: Ministerio, Instituto o Autoridad sanitaria) aparece en el detalle de cada partida."
            " Busca la partida y en su ficha verás la entidad responsable y los requisitos asociados."
        )
        sugerencias = ["Buscar partida", "Ver detalle de partida"]

    elif "tienen soporte" in m or "hay soporte" in m or "disponen de soporte" in m:
        respuesta = (
            "Sí, tenemos soporte. Puedes acceder al formulario de soporte desde el menú o en la página de Soporte."
            " Si es un problema urgente, envía los detalles y tu contacto para que el equipo te responda."
        )
        sugerencias = ["Ir a soporte", "¿Cuánto tardan en responder?"]

    elif "cuánto tardan en responder" in m:
        respuesta = (
            "Normalmente el equipo de soporte responde en menos de 24 horas hábiles. Si necesitas respuesta más rápida indícalo en tu solicitud y deja un teléfono o correo."
        )
        sugerencias = ["Contactar soporte", "¿Atienden por WhatsApp?"]

    elif re.search(r"\bice\b", m) or re.search(r"\biehd\b", m):
        respuesta = (
            "ICE (Impuesto al Consumo Específico) e IEHD (Impuesto Especial a Hidrocarburos y Derivados) son gravámenes que aplican a determinados productos (bebidas, tabacos, combustibles, etc.)."
            " En la ficha de la partida verás si aplica ese impuesto y su monto."
        )
        sugerencias = ["Buscar partida", "¿A qué partidas aplica?"]

    elif "requisitos de una partida" in m or "documentos de una partida" in m:
        respuesta = (
            "Revisa la ficha de la partida desde el buscador: allí se listan los documentos requeridos, la entidad emisora y la base legal."
        )
        sugerencias = ["Buscar partida", "Ver detalle de partida"]

    elif "capítulo arancelario" in m or "qué es un capítulo" in m:
        respuesta = (
            "Un capítulo arancelario agrupa partidas relacionadas por familia de productos. Por ejemplo, Capítulo 01: animales vivos; Capítulo 02: carnes."
        )
        sugerencias = ["Lista de capítulos", "Buscar partida"]

    elif "qué es una partida" in m or "partida arancelaria" in m:
        respuesta = (
            "Una partida arancelaria es un código que identifica un producto para efectos de impuesto y regulación en comercio internacional."
            " En su detalle se indican gravámenes, permisos y documentos asociados."
        )
        sugerencias = ["¿Qué es una subpartida?", "Buscar partida"]

    elif "buscar" in m or m == "buscar partida":
        respuesta = (
            """Cómo buscar una partida:
• Ve a 'Buscar Partidas' en el menú.
• Escribe el código completo o parte de la descripción.
• Usa filtros (capítulo, gravamen, entidad) para acotar resultados."""
        )
        sugerencias = ["Buscar partida", "Filtrar por capítulo"]

    elif "filtro" in m:
        respuesta = (
            """Filtros útiles:
• Capítulo: agrupa por familias.
• Gravamen: muestra el impuesto aplicable.
• Entidad que emite: autoridad responsable.
• Tipo de documento: qué se exige para el trámite."""
        )
        sugerencias = ["¿Qué filtro me conviene usar?", "Buscar partida"]

    elif "licencia" in m or "caduca" in m or m == "mi licencia" or "tiempo" in m or "cuánto tiempo" in m:
        # consultar licencia solo si el usuario está autenticado
        if not request.user.is_authenticated:
            respuesta = "Para ver el estado de tu licencia debes iniciar sesión. Pulsa en 'Iniciar sesión' y luego pregunta 'Mi licencia'."
            sugerencias = ["Iniciar sesión", "Crear cuenta"]
        else:
            try:
                from datetime import date
                from .models import LicenciaTemporal
                licencia = LicenciaTemporal.objects.filter(usuario=request.user, estado=True).order_by('-fecha_fin').first()
                if licencia:
                    hoy = date.today()
                    dias_restantes = (licencia.fecha_fin - hoy).days
                    if dias_restantes > 0:
                        respuesta = f"Tu licencia temporal vence el <strong>{licencia.fecha_fin}</strong>. Te quedan <strong>{dias_restantes} días</strong> de acceso al sistema. Si quieres renovarla, escribe '¿Cómo renovarla?'."
                    else:
                        respuesta = f"Tu licencia temporal expiró el <strong>{licencia.fecha_fin}</strong>. Para renovarla contacta al administrador o solicita soporte desde la sección 'Soporte'."
                else:
                    respuesta = "No se encontró una licencia activa para tu usuario. Contacta al administrador o abre un ticket en Soporte."
            except Exception:
                respuesta = "No se pudo obtener información de licencia en este momento. Intenta nuevamente más tarde."
            sugerencias = ["¿Qué es la licencia?", "Cómo renovarla"]

    elif "qué es la licencia" in m:
        respuesta = (
            "Una licencia es el permiso que autoriza tu acceso al sistema durante un periodo. Las cuentas nuevas suelen tener 7 días de prueba."
        )
        sugerencias = ["¿Está activa mi licencia?", "Cómo renovarla"]

    elif "manual buscador" in m or "manual del buscador" in m:
        respuesta = (
            "Puedes revisar el manual llamado <strong>Guía del Buscador</strong> en la sección de manuales.<br>"
            "Ahí se explica cómo hacer búsquedas por texto, código y filtros."
        )
        sugerencias = ["Ver manuales", "Buscar partida"]

    elif "manual administrador" in m or "manual para admin" in m:
        respuesta = (
            "Sí: en 'Manuales' hay guías para administradores (importar Excel, gestionar partidas y usuarios)."
        )
        sugerencias = ["Ver manuales", "¿Cómo importar partidas?"]

    elif "manual" in m or m == "ver manuales":
        respuesta = (
            "Puedes encontrar los manuales en la sección 'Manuales' del menú. Busca 'Guía del Buscador' o 'Manual Administrador' según lo que necesites."
        )
        sugerencias = ["Ver manuales", "Guía del Buscador"]

    elif "registro" in m or "crear cuenta" in m:
        respuesta = (
            "Para crear una cuenta ve a 'Crear cuenta' en la pantalla principal. Necesitarás nombre, apellido, correo, usuario y contraseña."
        )
        sugerencias = ["¿Qué datos necesito?", "Crear cuenta"]

    elif "qué datos necesito" in m or "datos para registrarme" in m:
        respuesta = (
            "Solicitamos: nombre, apellido, correo electrónico válido, nombre de usuario y contraseña segura."
        )
        sugerencias = ["Crear cuenta", "¿Puedo registrarme sin correo?"]

    elif "registrarme sin correo" in m:
        respuesta = (
            "No, actualmente es necesario un correo válido para completar el registro y recibir notificaciones."
        )
        sugerencias = ["Crear cuenta", "¿Qué datos necesito?"]

    elif "error" in m or "problema" in m or "no puedo" in m:
        respuesta = (
            """Lamento que tengas un problema. Prueba esto:
• Verifica usuario/contraseña.
• Revisa que tu licencia esté activa.
• Limpia caché o intenta otro navegador.
Si sigue fallando, abre un ticket en Soporte indicando pasos para reproducir el error."""
        )
        sugerencias = ["Contactar soporte", "No encuentro partidas", "No puedo iniciar sesión"]

    elif "soporte" in m or "contactar" in m or "ayuda" in m or m == "contactar soporte":
        respuesta = (
            "Para contactar soporte usa el formulario en 'Soporte' o envía un correo con tu problema y datos de contacto."
            " Si quieres, puedo abrir la página de Soporte ahora mismo para que completes el formulario."
        )
        # sugerencia inmediata para preguntar Whatsapp
        sugerencias = ["Ir a soporte", "¿Atienden por WhatsApp?"]
        # si el usuario pidió explícitamente abrir soporte (incluye frases como "contactar soporte"), devolver la acción inmediata
        if any(kw in m for kw in ["abre", "abrir", "abrí", "abre la pagina", "abre la página", "abrir soporte", "abre soporte"]) or 'contactar soporte' in m or m.strip() in ('contactar soporte','contactar'):
            action = {'open_support': '/soporte/', 'action_text': 'Abrir la página de Soporte.'}
            # ejecutar la acción automáticamente en el cliente
            action['auto_open'] = True
            respuesta = "Abriendo la página de Soporte..."
            return respuesta, sugerencias, action
        # en caso contrario, sugerimos la acción y la guardamos en session para que
        # si el usuario confirma con 'si' en el siguiente mensaje, podamos ejecutarla
        try:
            request.session['chat_last_action'] = {'open_support': '/soporte/', 'action_text': 'Abrir la página de Soporte.'}
        except Exception:
            pass
        return respuesta, sugerencias, None

    else:
        respuesta = (
            "No entendí exactamente tu consulta. ¿Quieres que te muestre opciones de ayuda? Puedes elegir: Buscar partida, Ver manuales, Mi licencia o Contactar soporte."
        )
        sugerencias = ["Buscar partida", "Ver manuales", "Mi licencia", "Contactar soporte"]

    return respuesta, sugerencias, action

# API endpoint para el chat de ayuda usando Dialogflow (si está configurado). Si no,
# se usa el generador local como fallback.
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt



@csrf_exempt
@require_POST
def api_chat_help(request):
    """Recibe POST con 'message' y devuelve JSON con respuesta del asistente SISARM.
    
    Parámetros:
      - message: Pregunta del usuario (requerido)
      - session_id: ID de sesión para mantener contexto (opcional)
      - public=1: Permite acceso anónimo (default: requiere login)
    
    Respuesta JSON:
      {
        "ok": true/false,
        "reply": "texto de respuesta",
        "action": null o {"open_support": "/soporte/"},
        "action_text": null o "texto del botón"
      }
    """
    
    # Determinar si la petición solicita acceso público (soporta form, querystring y JSON).
    # Por UX: permitir acceso anónimo por defecto para la página de ayuda/chat.
    is_public = False
    try:
        if request.POST.get('public') == '1' or request.GET.get('public') == '1':
            is_public = True
        else:
            import json as _json
            payload = _json.loads(request.body.decode('utf-8') or '{}')
            if str(payload.get('public')) == '1':
                is_public = True
    except Exception:
        # Si hay cualquier error al parsear, conservamos is_public=False y permitimos acceso anónimo
        is_public = False

    # Permitir acceso anónimo por defecto (este endpoint ofrece información pública de ayuda).
    if not request.user.is_authenticated and not is_public:
        is_public = True

    # Obtener el mensaje (soportar JSON y form-encoded)
    message = request.POST.get('message', '').strip()
    if not message:
        try:
            import json
            payload = json.loads(request.body.decode('utf-8') or '{}')
            message = payload.get('message', '').strip()
        except Exception:
            pass

    if not message:
        return JsonResponse({'ok': False, 'error': 'message es requerido'}, status=400)

    # Obtener session_id (para mantener contexto en Dialogflow)
    try:
        session_id = request.POST.get('session_id') or ''
        if not session_id:
            try:
                import json as _json
                payload = _json.loads(request.body.decode('utf-8') or '{}')
                session_id = payload.get('session_id', '')
            except Exception:
                pass
        
        if not session_id:
            # Generar session_id según usuario/sesión anónima
            if request.user.is_authenticated:
                session_id = f"user-{request.user.id}"
            else:
                sk = request.session.session_key or 'no-session'
                session_id = f"anon-{sk}"
    except Exception:
        session_id = 'default'

    # Registrar en historial de chat (si existe modelo ChatMessage)
    try:
        from .models import ChatMessage
        chat_entry = ChatMessage.objects.create(
            usuario=request.user if request.user.is_authenticated else None,
            mensaje=message,
            publico=is_public
        )
    except Exception:
        chat_entry = None

    # Obtener respuesta usando la función local mejorada
    reply = None
    sugerencias = []
    action = None
    try:
        # Usar la función local mejorada que devuelve (respuesta, sugerencias, action)
        reply, sugerencias, action = _generate_local_reply(message, request)
    except Exception as e:
        # Si hay error, usar respuesta por defecto
        reply = "No entendí tu pregunta. ¿Puedo ayudarte con algo más?"
        sugerencias = ["Buscar partida", "Ver manuales", "Contactar soporte"]
        action = None

    if not reply:
        reply = "No entendí exactamente. ¿Cuál es tu pregunta?"

    # Actualizar historial si existe
    if chat_entry:
        try:
            chat_entry.respuesta = reply
            chat_entry.save()
        except Exception:
            pass

    # Registrar en actividad del usuario
    try:
        from .models import HistoriaActividad
        HistoriaActividad.objects.create(
            usuario=request.user if request.user.is_authenticated else None,
            accion=f"chat_help: {message[:100]}"
        )
    except Exception:
        pass

    # Devolver respuesta en JSON con sugerencias y acciones
    return JsonResponse({
        'ok': True,
        'reply': reply,
        'sugerencias': sugerencias,
        'action': action
    })

# Vista para estadísticas de aranceles
@login_required
def estadisticas_aranceles(request):
    # filtros posibles: entidad_emite, capitulo (coma-separados)
    entidad = request.GET.get('entidad_emite', '').strip()
    capitulos_param = request.GET.get('capitulos', '').strip()
    desde = request.GET.get('desde', '').strip()
    hasta = request.GET.get('hasta', '').strip()

    qs = PartidaArancelaria.objects.all()
    if entidad:
        qs = qs.filter(entidad_emite__icontains=entidad)
    if capitulos_param:
        # aceptar lista separada por comas
        caps = [c.strip() for c in capitulos_param.split(',') if c.strip()]
        qs = qs.filter(capitulo__in=caps)

    # Nota: el modelo PartidaArancelaria no tiene campo de fecha de actualización ('updated_at'),
    # por lo que los filtros 'desde'/'hasta' no se aplican. Para habilitarlos es necesario añadir
    # un campo DateTimeField en el modelo y ejecutar migraciones. Informar al template si se pasan.
    fecha_filter_aplicada = False
    if desde or hasta:
        # intentar parsear fechas en formato YYYY-MM-DD y aplicar filtro sobre updated_at
        try:
            if desde:
                desde_dt = datetime.strptime(desde, '%Y-%m-%d')
                qs = qs.filter(updated_at__gte=desde_dt)
            if hasta:
                hasta_dt = datetime.strptime(hasta, '%Y-%m-%d')
                # incluir fin del día
                hasta_dt = hasta_dt.replace(hour=23, minute=59, second=59)
                qs = qs.filter(updated_at__lte=hasta_dt)
            fecha_filter_aplicada = True
        except Exception:
            # si el parse falla, no aplicamos filtro y marcamos que no se aplicó
            fecha_filter_aplicada = False

    # calcular promedios por capítulo: parsear el campo 'gravamen' con tolerancia a textos
    from decimal import Decimal, InvalidOperation
    import json as _json

    def parse_gravamen(val):
        if not val:
            return None
        s = str(val).strip()
        # buscar primera ocurrencia de número (acepta coma como decimal)
        m = re.search(r"[-+]?\d+[\.,]?\d*", s)
        if not m:
            return None
        num = m.group(0).replace(',', '.')
        try:
            return Decimal(num)
        except InvalidOperation:
            return None

    capitulos = list(qs.values_list('capitulo', flat=True).distinct())
    promedios = []
    for cap in capitulos:
        sub_qs = qs.filter(capitulo=cap)
        total = Decimal('0')
        count = 0
        values = []
        for p in sub_qs:
            g = parse_gravamen(p.gravamen)
            if g is None:
                continue
            values.append(g)
            total += g
            count += 1
        promedio = (total / count) if count > 0 else None
        # calcular min y max si tenemos valores
        min_v = (min(values) if values else None)
        max_v = (max(values) if values else None)
        promedios.append({
            'capitulo': cap,
            'promedio': (round(promedio, 2) if promedio is not None else None),
            'min': (round(min_v, 2) if min_v is not None else None),
            'max': (round(max_v, 2) if max_v is not None else None),
            'count': count
        })

    # ordenar por capítulo (alfanumérico)
    promedios = sorted(promedios, key=lambda x: (x['capitulo'] or ''))

    # pasar JSON para el gráfico
    promedios_json = _json.dumps(promedios, default=str)

    return render(request, 'partidas/estadisticas_aranceles.html', {
        'promedios': promedios,
        'promedios_json': promedios_json,
        'filtros': {
            'entidad_emite': entidad,
            'capitulos': capitulos_param,
            'desde': desde,
            'hasta': hasta,
            'fecha_filter_aplicada': fecha_filter_aplicada
        }
    })


@login_required
def log_click(request):
    """Endpoint para registrar clicks desde la UI (historial -> detalle / búsqueda).
    Espera POST con 'partida_id' (opcional) y 'accion' (opcional) y 'termino' (opcional).
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required'}, status=400)

    partida_id = request.POST.get('partida_id') or request.POST.get('partida')
    accion = request.POST.get('accion', 'historial_click')
    termino = request.POST.get('termino', '')

    partida = None
    if partida_id:
        try:
            partida = PartidaArancelaria.objects.get(id=partida_id)
        except PartidaArancelaria.DoesNotExist:
            partida = None

    try:
        ClickLog.objects.create(usuario=request.user, partida=partida, accion=accion, extra=termino)
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@rol_requerido('Administrador')
def admin_busquedas(request):
    """Vista para que Admin vea, filtre y exporte búsquedas realizadas por despachantes."""
    qs = Busqueda.objects.select_related('usuario').order_by('-fecha')

    usuario = request.GET.get('usuario', '').strip()
    termino = request.GET.get('termino', '').strip()
    desde = request.GET.get('desde', '').strip()
    hasta = request.GET.get('hasta', '').strip()

    if usuario:
        qs = qs.filter(usuario__username__icontains=usuario)
    if termino:
        qs = qs.filter(termino_buscado__icontains=termino)
    if desde:
        qs = qs.filter(fecha__gte=desde)
    if hasta:
        qs = qs.filter(fecha__lte=hasta)

    # Export CSV
    if request.GET.get('export') == 'csv':
        # registrar la exportación
        filtros = { 'usuario': usuario, 'termino': termino, 'desde': desde, 'hasta': hasta }
        ExportLog.objects.create(usuario=request.user, accion='export_busquedas_csv', filtros=str(filtros))

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="busquedas_export.csv"'
        writer = csv.writer(response)
        writer.writerow(['usuario', 'termino', 'fecha', 'resultados'])
        for b in qs:
            resultados_text = getattr(b, 'resultados', '') or ''
            writer.writerow([b.usuario.username, b.termino_buscado, b.fecha.strftime('%Y-%m-%d %H:%M:%S'), resultados_text])
        return response

    # paginar manualmente simple: mostrar primeras 500
    resultados = qs[:500]

    return render(request, 'partidas/admin_busquedas.html', {'busquedas': resultados, 'filtros': {'usuario': usuario, 'termino': termino, 'desde': desde, 'hasta': hasta}})


@login_required
def solicitar_ayuda_documento(request, referencia_id):
    """Permite al usuario solicitar ayuda cuando un documento no está disponible.
    Crea una entrada en HistoriaActividad y redirige de vuelta al detalle de la partida.
    """
    from django.shortcuts import redirect
    ref = None
    try:
        ref = PartidaReferencia.objects.get(id=referencia_id)
    except PartidaReferencia.DoesNotExist:
        messages.error(request, 'Referencia no encontrada')
        return redirect('inicio')

    # crear registro de actividad solicitando ayuda
    mensaje = f"Solicitud de ayuda para documento {ref.id} (partida {ref.partida.codigo}) por usuario {request.user.username}"
    HistoriaActividad.objects.create(usuario=request.user, accion=mensaje)
    messages.success(request, 'Se ha solicitado ayuda. El equipo de soporte recibirá la solicitud.')
    return redirect('detalle_partida', partida_id=ref.partida.id)


# AJAX endpoints para el panel de estadísticas en admin
@login_required
@rol_requerido('Administrador')
def api_stats_by_chapter(request):
    """API AJAX para obtener estadísticas de gravamen por capítulo con filtros.
    GET params: entidad_emite, capitulos (coma-separados)
    Retorna JSON con lista de estadísticas.
    """
    from django.http import JsonResponse
    from decimal import Decimal, InvalidOperation
    import re
    
    entidad = request.GET.get('entidad_emite', '').strip()
    capitulos_param = request.GET.get('capitulos', '').strip()
    
    try:
        def _parse_gravamen(val):
            if not val:
                return None
            s = str(val).strip()
            m = re.search(r"[-+]?\d+[\.,]?\d*", s)
            if not m:
                return None
            num = m.group(0).replace(',', '.')
            try:
                return Decimal(num)
            except InvalidOperation:
                return None
        
        part_qs = PartidaArancelaria.objects.all()
        if entidad:
            part_qs = part_qs.filter(entidad_emite__icontains=entidad)
        if capitulos_param:
            caps = [c.strip() for c in capitulos_param.split(',') if c.strip()]
            part_qs = part_qs.filter(capitulo__in=caps)
        
        chap_map = {}
        for p in part_qs:
            cap = p.capitulo or 'Sin capítulo'
            g = _parse_gravamen(p.gravamen)
            if g is None:
                continue
            data = chap_map.setdefault(cap, {'values': [], 'count': 0})
            data['values'].append(g)
            data['count'] += 1
        
        chapter_stats = []
        for cap, data in chap_map.items():
            vals = data['values']
            if not vals:
                continue
            mn = min(vals)
            mx = max(vals)
            tot = sum(vals)
            avg = (tot / len(vals)) if len(vals) > 0 else None
            chapter_stats.append({
                'capitulo': cap,
                'promedio': round(float(avg), 2) if avg is not None else None,
                'min': round(float(mn), 2) if mn is not None else None,
                'max': round(float(mx), 2) if mx is not None else None,
                'count': data['count']
            })
        chapter_stats = sorted(chapter_stats, key=lambda x: (x['capitulo'] or ''))
        return JsonResponse({'success': True, 'data': chapter_stats, 'count': len(chapter_stats)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@rol_requerido('Administrador')
def api_autocomplete_entidades(request):
    """API para autocompletado de entidades emisoras.
    GET param: q (búsqueda parcial)
    Retorna JSON con lista de entidades únicas.
    """
    from django.http import JsonResponse
    q = request.GET.get('q', '').strip().lower()
    # Obtener lista única de entidades
    entidades = PartidaArancelaria.objects.values_list('entidad_emite', flat=True).distinct().order_by('entidad_emite')
    # Filtrar por búsqueda
    if q:
        entidades = [e for e in entidades if q in e.lower()]
    else:
        entidades = list(entidades)[:30]  # Límite para no sobrecargar
    return JsonResponse({'results': entidades})


@login_required
@rol_requerido('Administrador')
def api_autocomplete_capitulos(request):
    """API para autocompletado de capítulos.
    GET param: q (búsqueda parcial)
    Retorna JSON con lista de capítulos únicos.
    """
    from django.http import JsonResponse
    q = request.GET.get('q', '').strip().lower()
    # Obtener lista única de capítulos
    capitulos = PartidaArancelaria.objects.values_list('capitulo', flat=True).distinct().order_by('capitulo')
    # Filtrar por búsqueda
    if q:
        capitulos = [c for c in capitulos if q in c.lower()]
    else:
        capitulos = list(capitulos)[:30]
    return JsonResponse({'results': capitulos})

